import streamlit as st
import pandas as pd

st.set_page_config(page_title="BOM工具集", layout="wide")

def bomTools():
    s = st.sidebar.radio('BOM工具集',
                         (#'BOM数据库维护',
                          #'数据库问题核查',
                          'BOM数据表核对',
                          #'BOM数据库核对',
                          #'BOM差异件核对',
                          '差异件清单生成',
                          'MBOM差异件清单生成',
                          'EBOM&MBOM差异件清单生成',
                          'LOU核查工具',
                          'LOU打点工具',
                          #'工程配置工具',
                          'CMAN统计工具',
                          '配置表处理工具',
                          '配置打点信息一致性核查',
                          ))

    def BomLineStr(bomline):
        r = ''
        for x in bomline:
            if r == '':
                r = str(x)
            else:
                r = r + ' - ' + str(x)
        return r

    def checkTreeData(td1, td2):
        keys = ['id', 'name', 'count', 'level']
        b = True
        for k in keys:
            if not(td1[k] == td2[k]):
                b = False
                break
        return b

    def compare_trees(tree1, tree2):
        diff = []
        #if tree1["data"] != tree2["data"]:
        if not checkTreeData(tree1['data'], tree2['data']):
            diff.append({"id": tree1["data"]["id"], "attribute": "data", "tree1": tree1["data"], "tree2": tree2["data"]})
        children1 = {}
        children2 = {}
        for child1 in tree1['children']:
            children1[child1['data']['name']] = child1
        for child2 in tree2['children']:
            children2[child2['data']['name']] = child2

        comparedChildIds = []
        #st.write(children1.keys())
        #st.write(children2.keys())
        for cid in children1.keys():
            if cid not in comparedChildIds:
                comparedChildIds.append(cid)
                child1 = children1[cid]
                if cid in children2.keys():
                    subtree_diff = compare_trees(children1[cid], children2[cid])
                    diff.extend(subtree_diff)
                else:
                    diff.append({"id": cid, "attribute": "children", "tree1": child1['data'], "tree2": None})
        for cid in children2.keys():
            if cid not in comparedChildIds:
                comparedChildIds.append(cid)
                child2 = children2[cid]
                if cid in children1.keys():
                    subtree_diff = compare_trees(children1[cid], children2[cid])
                    diff.extend(subtree_diff)
                else:
                    diff.append({"id": cid, "attribute": "children", "tree1": None, "tree2": child2['data']}) 

        
        #for i in range(max(len(tree1["children"]), len(tree2["children"]))):
        #    if i < len(tree1["children"]) and i < len(tree2["children"]):
        #        subtree_diff = compare_trees(tree1["children"][i], tree2["children"][i])
        #        diff.extend(subtree_diff)
        #    elif i < len(tree1["children"]):
        #        diff.append({"id": tree1["children"][i]["data"]["id"], "attribute": "children", "tree1": tree1["children"][i]['data'], "tree2": None})
        #    else:
        #        diff.append({"id": tree2["children"][i]["data"]["id"], "attribute": "children", "tree1": None, "tree2": tree2["children"][i]['data']})

        return diff

    def SystemChildrenDiff_DB(ss, cols1, cols2, container):
        s0 = ss['sys']
        sid0 = s0[0]
        sname0 = s0[1]
        sc0 = ss['children']

        sys_db = TinyDB('sys_db.json')
        Unit = Query()
        rs = sys_db.search((Unit.id == sid0) | (Unit.parent == sid0))
        if len(rs) <= 1:
            cols1[0].write('数据库中没有【%s %s】这个2Y系统' % (sid0, sname0))
            return

        dbS1 = rs[0]
        dbSc1 = rs[1:]
        s1 = [dbS1['id'], dbS1['name'],dbS1['count'],dbS1['level']]
        sc1 = []
        for unit in dbSc1:
            sc1.append([unit['id'], unit['name'],unit['count'],unit['level']])
            
        r = False

        n0 = len(sc0)
        n1 = len(sc1)
        #if n0 == n1:
        #    for i in range(0,n0):
        #        b0 = sc0[i]
        #        b1 = sc1[i]
        #        if (b0[0] == b1[0]) and not(str(b0[1]) == str(b1[1]) and str(b0[2]) == str(b1[2]) and str(b0[3]) == str(b1[3])):
        #            r = True
        #            break

        ###########
        for i in range(0,n0):
            b0 = sc0[i]
            id0 = b0[0]            
            bIn = False
            for j in range(0, n1):                
                b1 = sc1[j]
                id1 = b1[0]
                if (id0 == id1):
                    bIn = True
                    r =  not(str(b0[1]) == str(b1[1]) and str(b0[2]) == str(b1[2]) and str(b0[3]) == str(b1[3]))
                    if not(r):
                        break
            r = r or (not bIn)
            if r:
                break 

        if not r:
            for i in range(0,n1):
                b1 = sc1[i]
                id1 = b1[0]            
                bIn = False
                for j in range(0, n0):                
                    b0 = sc0[j]
                    id0 = b0[0]
                    if (id0 == id1):
                        bIn = True
                        r =  not(str(b0[1]) == str(b1[1]) and str(b0[2]) == str(b1[2]) and str(b0[3]) == str(b1[3]))
                        if not(r):
                            break
                r = r or (not bIn)
                if r:
                    break
        ###########

        diff2ys = []
        if r:
            if (s0[1] == s1[1]):
                cols2[0].markdown(':blue['+BomLineStr(s0)+']')
                cols2[1].markdown(':blue['+BomLineStr(s1)+']')
            else:
                cols2[0].markdown(':red['+BomLineStr(s0)+']')
                cols2[1].markdown(':red['+BomLineStr(s1)+']')

            
            ##############################
            dict0 = {}
            dict1 = {}
            
            d0s = []
            u0 = {'id':s0[0], 'name':s0[1], 'count':str(s0[2]), 'level':s0[3]}
            d0s.append(u0)
            dict0[s0[0]] = u0
            for d in sc0:
                u0 = {'id':d[0], 'name':d[1], 'count':str(d[2]), 'level':d[3]}
                d0s.append(u0)
                dict0[d[0]] = u0
            d0 = pd.DataFrame(d0s)

            d1s = []
            u1 = {'id':s1[0], 'name':s1[1], 'count':str(s1[2]), 'level':s1[3]}
            d1s.append(u1)
            dict1[s1[0]] = u1
            for d in sc1:
                u1 = {'id':d[0], 'name':d[1], 'count':str(d[2]), 'level':d[3]}
                d1s.append(u1)
                dict1[d[0]] = u1
            d1 = pd.DataFrame(d1s)

            td = [s0]+sc0
            td = [{'id':s0[0], 'name':s0[1], 'count':str(s0[2]), 'level':s0[3]}]
            for c0 in sc0:
                td.append({'id':c0[0], 'name':c0[1], 'count':str(c0[2]), 'level':c0[3]})
            tree0 = generate_tree(solveTreeData(td), 0)

            td = [s1]+sc1
            td = [{'id':s1[0], 'name':s1[1], 'count':str(s1[2]), 'level':s1[3]}]
            for c1 in sc1:
                td.append({'id':c1[0], 'name':c1[1], 'count':str(c1[2]), 'level':c1[3]})
            tree1 = generate_tree(solveTreeData(td), 0)

            ids0 = dict0.keys()
            ids1 = dict1.keys()                   

            diffs = compare_trees(tree0[0], tree1[0])
            if diffs:               
                rds0 = []
                rds1 = []
                rds = []
                eds = []
                nullUnit = {'id':None, 'name':None, 'count':None, 'level':None}

                for d0 in d0s:
                    id0 = d0['id']
                    u0 = dict0[id0]
                    if (id0 in ids0) and (id0 in ids1):
                        if not (id0 in eds):                            
                            u1 = dict1[id0]
                            rds0.append(dict0[id0])
                            rds1.append(dict1[id0])
                            rds.append({'id0':u0['id'],'name0':u0['name'],'count0':u0['count'],'level0':u0['level'], 'id1':u1['id'],'name1':u1['name'],'count1':u1['count'],'level1':u1['level']})
                            eds.append(id0)
                    else:
                        rds0.append(dict0[id0])                        
                        rds1.append(nullUnit)
                        rds.append({'id0':u0['id'],'name0':u0['name'],'count0':u0['count'],'level0':u0['level'], 'id1':None,'name1':None,'count1':None,'level1':None})
                for d1 in d1s:
                    id1 = d1['id']
                    u1 = dict1[id1]
                    if (id1 in ids0) and (id1 in ids1):
                        if not (id1 in eds):
                            u0 = dict0[id1]
                            rds0.append(dict0[id1])
                            rds1.append(dict1[id1])
                            rds.append({'id0':u0['id'],'name0':u0['name'],'count0':u0['count'],'level0':u0['level'], 'id1':u1['id'],'name1':u1['name'],'count1':u1['count'],'level1':u1['level']})
                            eds.append(id1)
                    else:
                        rds0.append(nullUnit)
                        rds1.append(dict1[id1])
                        rds.append( {'id0':None,'name0':None,'count0':None,'level0':None, 'id1':u1['id'],'name1':u1['name'],'count1':u1['count'],'level1':u1['level']})
                d0 = pd.DataFrame(rds0)
                d1 = pd.DataFrame(rds1)                
                color = ((d0.name != d1.name) | (d0['count'] != d1['count']) | (d0.level != d1.level)).map({True: 'background-color: yellow', False: ''})
                d0 = d0.style.apply(lambda s: color)
                d1 = d1.style.apply(lambda s: color)
                d = pd.DataFrame(rds)
                color = ((d.name0 != d.name1) | (d.count0 != d.count1) | (d.level0 != d.level1)).map({True: 'background-color: yellow', False: ''})
                d = d.style.apply(lambda s: color)
                container.dataframe(d, use_container_width=True)                          
                        
                diffDatas0 = [] #[{'零件编号':s0[0], '零件名称':s0[1], '数量':s0[2], '层级':s0[3]}]
                diffDatas1 = [] #[{'零件编号':s1[0], '零件名称':s1[1], '数量':s1[2], '层级':s1[3]}]
                for diff in diffs:
                    t1 = diff['tree1']
                    t2 = diff['tree2']
                    diffData0 = {'零件编号':None, '零件名称':None, '数量':None, '层级':None}
                    diffData1 = {'零件编号':None, '零件名称':None, '数量':None, '层级':None}
                    if t1:
                        diffData0['零件编号'] = t1['id']
                        diffData0['零件名称'] = t1['name']
                        diffData0['数量'] = t1['count']
                        diffData0['层级'] = t1['level']
                    if t2:
                        diffData1['零件编号'] = t2['id']
                        diffData1['零件名称'] = t2['name']
                        diffData1['数量'] = t2['count']
                        diffData1['层级'] = t2['level']
                    diffDatas0.append(diffData0)
                    diffDatas1.append(diffData1)
                cols2 = container.columns(2)
                cols2[0].dataframe(diffDatas0, use_container_width=True)
                cols2[1].dataframe(diffDatas1, use_container_width=True)            
            ##############################
            if 0:
                d0s = []
                d0s.append({'id':s0[0], 'name':s0[1], 'count':s0[2], 'level':s0[3]})
                for d in sc0:
                    d0s.append({'id':d[0], 'name':d[1], 'count':d[2], 'level':d[3]})
                d0 = pd.DataFrame(d0s)

                d1s = []
                d1s.append({'id':s1[0], 'name':s1[1], 'count':s1[2], 'level':s1[3]})
                for d in sc1:
                    d1s.append({'id':d[0], 'name':d[1], 'count':d[2], 'level':d[3]})
                d1 = pd.DataFrame(d1s)
                d0
                d1
                color = ((d0.name != d1.name) | (d0.level != d1.level)| (d0['count'] != d1['count'])).map({True: 'background-color: yellow', False: ''})
                d0 = d0.style.apply(lambda s: color)
                d1 = d1.style.apply(lambda s: color)
                cols2[0].dataframe(d0, use_container_width=True)
                #cols2[1].markdown(':red['+BomLineStr(b1)+']')
                cols2[1].dataframe(d1, use_container_width=True)


            td = [s0]+sc0
            td = [{'id':s0[0], 'name':s0[1], 'count':s0[2], 'level':s0[3]}]
            for c0 in sc0:
                td.append({'id':c0[0], 'name':c0[1], 'count':c0[2], 'level':c0[3]})
            tree0 = generate_tree(solveTreeData(td), 0)

            td = [s1]+sc1
            td = [{'id':s1[0], 'name':s1[1], 'count':s1[2], 'level':s1[3]}]
            for c1 in sc1:
                td.append({'id':c1[0], 'name':c1[1], 'count':c1[2], 'level':c1[3]})
            tree1 = generate_tree(solveTreeData(td), 0)

            if 0:
                diffs = compare_trees(tree0[0], tree1[0])
                if diffs:            
                    diffDatas0 = [] #[{'零件编号':s0[0], '零件名称':s0[1], '数量':s0[2], '层级':s0[3]}]
                    diffDatas1 = [] #[{'零件编号':s1[0], '零件名称':s1[1], '数量':s1[2], '层级':s1[3]}]
                    for diff in diffs:
                        t1 = diff['tree1']
                        t2 = diff['tree2']
                        diffData0 = {'零件编号':None, '零件名称':None, '数量':None, '层级':None}
                        diffData1 = {'零件编号':None, '零件名称':None, '数量':None, '层级':None}
                        if t1:
                            diffData0['零件编号'] = t1['id']
                            diffData0['零件名称'] = t1['name']
                            diffData0['数量'] = t1['count']
                            diffData0['层级'] = t1['level']
                        if t2:
                            diffData1['零件编号'] = t2['id']
                            diffData1['零件名称'] = t2['name']
                            diffData1['数量'] = t2['count']
                            diffData1['层级'] = t2['level']
                        diffDatas0.append(diffData0)
                        diffDatas1.append(diffData1) 
                    cols2[0].dataframe(diffDatas0, use_container_width=True)
                    cols2[1].dataframe(diffDatas1, use_container_width=True)
        else:
            if not (s0[1] == s1[1]):
                cols1[0].markdown(':red['+BomLineStr(s0)+']')
                cols1[1].markdown(':red['+BomLineStr(s1)+']') 

    def solveSystemChildren(sc):
        cnameCounts = {}
        tsc = []
        for c in sc:
            cid = c[0]
            cname = c[1]
            if cname in cnameCounts:
                cnameCount = cnameCounts[cname]
                c[1] = '%s_%d' % (c[1], cnameCount)
                cnameCounts[cname] = cnameCount + 1                
            else:
               cnameCounts[cname] = 1 
            tsc.append(c)
        return tsc            

    def SystemDiff(ss0, ss1, container):
        diffDatas = []
        with container:
            try:
                if ss0 and ss1:            
                    s0 = ss0['sys']
                    s1 = ss1['sys']
                    sc0 = ss0['children']
                    sc1 = ss1['children']
                    sc0 = solveSystemChildren(sc0)
                    sc1 = solveSystemChildren(sc1)
                    r = False
                    n0 = len(sc0)
                    n1 = len(sc1)
                    for i in range(0,n0):
                        b0 = sc0[i]
                        id0 = b0[0]            
                        bIn = False
                        for j in range(0, n1):                
                            b1 = sc1[j]
                            id1 = b1[0]
                            if (id0 == id1):
                                bIn = True
                                r =  not(str(b0[1]) == str(b1[1]) and str(b0[2]) == str(b1[2]) and str(b0[3]) == str(b1[3]))
                                if not(r):
                                    break
                        r = r or (not bIn)
                        if r:
                            break 
                    if not r:
                        for i in range(0,n1):
                            b1 = sc1[i]
                            id1 = b1[0]            
                            bIn = False
                            for j in range(0, n0):                
                                b0 = sc0[j]
                                id0 = b0[0]
                                if (id0 == id1):
                                    bIn = True
                                    r =  not(str(b0[1]) == str(b1[1]) and str(b0[2]) == str(b1[2]) and str(b0[3]) == str(b1[3]))
                                    if not(r):
                                        break
                            r = r or (not bIn)
                            if r:
                                break               
                    if r:
                        dict0 = {}
                        dict1 = {}
                        
                        d0s = []
                        u0 = {'id':s0[0], 'name':s0[1], 'count':str(s0[2]), 'level':s0[3], 'user':s0[4]}
                        d0s.append(u0)
                        dict0[s0[0]] = u0
                        for d in sc0:
                            u0 = {'id':d[0], 'name':d[1], 'count':str(d[2]), 'level':d[3], 'user':d[4]}
                            d0s.append(u0)
                            dict0[d[0]] = u0
                        d0 = pd.DataFrame(d0s)

                        d1s = []
                        u1 = {'id':s1[0], 'name':s1[1], 'count':str(s1[2]), 'level':s1[3], 'user':s1[4]}
                        d1s.append(u1)
                        dict1[s1[0]] = u1
                        for d in sc1:
                            u1 = {'id':d[0], 'name':d[1], 'count':str(d[2]), 'level':d[3], 'user':d[4]}
                            d1s.append(u1)
                            dict1[d[0]] = u1
                        d1 = pd.DataFrame(d1s)

                        td = [s0]+sc0
                        td = [{'id':s0[0], 'name':s0[1], 'count':str(s0[2]), 'level':s0[3], 'user':s0[4]}]
                        for c0 in sc0:
                            td.append({'id':c0[0], 'name':c0[1], 'count':str(c0[2]), 'level':c0[3], 'user':c0[4]})
                        tree0 = generate_tree(solveTreeData(td), 0)

                        td = [s1]+sc1
                        td = [{'id':s1[0], 'name':s1[1], 'count':str(s1[2]), 'level':s1[3], 'user':s1[4]}]
                        for c1 in sc1:
                            td.append({'id':c1[0], 'name':c1[1], 'count':str(c1[2]), 'level':c1[3], 'user':c1[4]})
                        tree1 = generate_tree(solveTreeData(td), 0)

                        ids0 = dict0.keys()
                        ids1 = dict1.keys()                   

                        diffs = compare_trees(tree0[0], tree1[0])                        
                        if diffs:
                            b2Y = False
                            for diff in diffs:
                                t1 = diff['tree1']
                                t2 = diff['tree2']
                                diffData = {'零件编号1':'/', '零件名称1':'/', '数量1':'/', '层级1':'/','工程师1':'/','备注1':'','零件编号2':'/', '零件名称2':'/', '数量2':'/', '层级2':'/','备注2':''}
                                if t1:
                                    if t1['level'] == '2Y':
                                        b2Y = True
                                    diffData['零件编号1'] = t1['id']
                                    diffData['零件名称1'] = t1['name']
                                    diffData['数量1'] = t1['count']
                                    diffData['层级1'] = t1['level']
                                    diffData['工程师1'] = t1['user']
                                    diffData['备注1'] = ''
                                if t2:
                                    if t2['level'] == '2Y':
                                        b2Y = True
                                    diffData['零件编号2'] = t2['id']
                                    diffData['零件名称2'] = t2['name']
                                    diffData['数量2'] = t2['count']
                                    diffData['层级2'] = t2['level']
                                    diffData['备注2'] = ''
                                diffDatas.append(diffData)
                            if not b2Y:
                                data2Y = {'零件编号1':s0[0], '零件名称1':s0[1], '数量1':s0[2], '层级1':s0[3],'工程师1':s0[4],'备注1':'','零件编号2':s1[0], '零件名称2':s1[1], '数量2':s1[2], '层级2':s1[3],'备注2':''}
                                diffDatas.insert(0, data2Y)
                                #'xxxxxxxxxxxxx'
                                #diffDatas
                            #container.dataframe(diffDatas, use_container_width=True)
                        else:
                            pass
                            #'Error: No diffs 1'
                    else:
                        pass
                        #'Error: No diffs 2' 
                elif ss0:
                    s0 = ss0['sys']
                    sc0 = ss0['children']                    
                    diffData = {'零件编号1':s0[0], '零件名称1':s0[1], '数量1':s0[2], '层级1':s0[3],'工程师1':s0[4],'备注1':'','零件编号2':'/', '零件名称2':'/', '数量2':'/', '层级2':'/','备注2':''}
                    diffDatas.append(diffData)
                    for c0 in sc0:
                        diffData = {'零件编号1':'/', '零件名称1':'/', '数量1':'/', '层级1':'/','工程师1':'/','备注1':'','零件编号2':'/', '零件名称2':'/', '数量2':'/', '层级2':'/','备注2':''}
                        diffData['零件编号1'] = c0[0]
                        diffData['零件名称1'] = c0[1]
                        diffData['数量1'] = c0[2]
                        diffData['层级1'] = c0[3]
                        diffData['工程师1'] = c0[4]
                        diffData['备注1'] = ''
                        diffDatas.append(diffData)
                elif ss1:   
                    s1 = ss1['sys']
                    sc1 = ss1['children']                    
                    diffData = {'零件编号1':'/', '零件名称1':'/', '数量1':'/', '层级1':'/','工程师1':'/','备注1':'','零件编号2':s1[0], '零件名称2':s1[1], '数量2':s1[2], '层级2':s1[3],'备注2':''}
                    diffDatas.append(diffData)
                    for c1 in sc1:
                        diffData = {'零件编号1':'/', '零件名称1':'/', '数量1':'/', '层级1':'/','工程师1':'/','备注1':'','零件编号2':'/', '零件名称2':'/', '数量2':'/', '层级2':'/','备注2':''}
                        diffData['零件编号2'] = c1[0]
                        diffData['零件名称2'] = c1[1]
                        diffData['数量2'] = c1[2]
                        diffData['层级2'] = c1[3]
                        diffData['备注2'] = ''
                        diffDatas.append(diffData)
            except Exception as e:
                'Error: 出错了！'
                #e
                col1, col2 = st.columns(2)
                with col1:
                    'ss0:',ss0
                with col2:
                    'ss1:',ss1
        return diffDatas
    
    def solveSystemChildren_MBOM(sc):
        cnameCounts = {}
        tsc = []
        for c in sc:
            cname = c[2]
            if cname in cnameCounts:
                cnameCount = cnameCounts[cname]
                c[2] = '%s_%d' % (c[2], cnameCount)
                cnameCounts[cname] = cnameCount + 1                
            else:
               cnameCounts[cname] = 1 
            tsc.append(c)
        return tsc
    
    def SystemDiff_MBOM(ss0, ss1, container):
        diffDatas = []
        with container:
            try:
                if ss0 and ss1:            
                    s0 = ss0['sys']
                    s1 = ss1['sys']
                    sc0 = ss0['children']
                    sc1 = ss1['children']
                    sc0 = solveSystemChildren_MBOM(sc0)
                    sc1 = solveSystemChildren_MBOM(sc1)
                    r = not(s0[1] == s1[1])
                    if not r:
                        n0 = len(sc0)
                        n1 = len(sc1)
                        for i in range(0, n0):
                            b0 = sc0[i]
                            id0 = b0[1]                                   
                            bIn = False
                            for j in range(0, n1):                
                                b1 = sc1[j]
                                id1 = b1[1]
                                if (id0 == id1):
                                    bIn = True                                
                                    r =  not(str(b0[0]) == str(b1[0]) and str(b0[2]) == str(b1[2]) and str(b0[3]) == str(b1[3]))
                                    if not(r):
                                        break
                            r = r or (not bIn)
                            if r:
                                break 
                        if not r:
                            for i in range(0,n1):
                                b1 = sc1[i]
                                id1 = b1[1]            
                                bIn = False
                                for j in range(0, n0):                
                                    b0 = sc0[j]
                                    id0 = b0[1]
                                    if (id0 == id1):
                                        bIn = True
                                        r =  not(str(b0[0]) == str(b1[0]) and str(b0[2]) == str(b1[2]) and str(b0[3]) == str(b1[3]))
                                        if not(r):
                                            break
                                r = r or (not bIn)
                                if r:
                                    break                              
                    if r:
                        dict0 = {}
                        dict1 = {}
                        
                        d0s = []
                        u0 = {'id':s0[1], 'name':s0[2], 'count':str(s0[3]), 'level':s0[0], 'user':''}                        
                        d0s.append(u0)
                        dict0[s0[0]] = u0
                        for d in sc0:
                            u0 = {'id':d[1], 'name':d[2], 'count':str(d[3]), 'level':d[0], 'user':''}
                            d0s.append(u0)
                            dict0[d[1]] = u0
                        d0 = pd.DataFrame(d0s)

                        d1s = []
                        u1 = {'id':s1[1], 'name':s1[2], 'count':str(s1[3]), 'level':s1[0], 'user':''}
                        d1s.append(u1)
                        dict1[s1[1]] = u1
                        for d in sc1:
                            u1 = {'id':d[1], 'name':d[2], 'count':str(d[3]), 'level':d[0], 'user':''}
                            d1s.append(u1)
                            dict1[d[1]] = u1
                        d1 = pd.DataFrame(d1s)

                        td = [s0]+sc0
                        td = [{'id':s0[1], 'name':s0[2], 'count':str(s0[3]), 'level':s0[0], 'user':''}]
                        for c0 in sc0:
                            td.append({'id':c0[1], 'name':c0[2], 'count':str(c0[3]), 'level':c0[0], 'user':''})
                        tree0 = generate_tree(solveTreeData_MBOM(td), 0)

                        td = [s1]+sc1
                        td = [{'id':s1[1], 'name':s1[2], 'count':str(s1[3]), 'level':s1[0], 'user':''}]
                        for c1 in sc1:
                            td.append({'id':c1[1], 'name':c1[2], 'count':str(c1[3]), 'level':c1[0], 'user':''})
                        tree1 = generate_tree(solveTreeData_MBOM(td), 0)

                        ids0 = dict0.keys()
                        ids1 = dict1.keys()                   

                        diffs = compare_trees(tree0[0], tree1[0])                        
                        if diffs:
                            b2Y = False
                            for diff in diffs:
                                t1 = diff['tree1']
                                t2 = diff['tree2']
                                diffData = {'零件编号1':'/', '零件名称1':'/', '数量1':'/', '层级1':'/','工程师1':'/','备注1':'','零件编号2':'/', '零件名称2':'/', '数量2':'/', '层级2':'/','备注2':''}
                                if t1:
                                    if t1['level'] == 1:
                                        b2Y = True
                                    diffData['零件编号1'] = t1['id']
                                    diffData['零件名称1'] = t1['name']
                                    diffData['数量1'] = t1['count']
                                    diffData['层级1'] = t1['level']
                                    diffData['工程师1'] = t1['user']
                                    diffData['备注1'] = ''
                                if t2:
                                    if t2['level'] == '2Y':
                                        b2Y = True
                                    diffData['零件编号2'] = t2['id']
                                    diffData['零件名称2'] = t2['name']
                                    diffData['数量2'] = t2['count']
                                    diffData['层级2'] = t2['level']
                                    diffData['备注2'] = ''
                                diffDatas.append(diffData)
                            if not b2Y:
                                data2Y = {'零件编号1':s0[1], '零件名称1':s0[2], '数量1':s0[3], '层级1':s0[0],'工程师1':'','备注1':'','零件编号2':s1[1], '零件名称2':s1[2], '数量2':s1[3], '层级2':s1[0],'备注2':''}
                                diffDatas.insert(0, data2Y)
                        else:
                            pass
                            #'Error: No diffs 1'
                    else:
                        pass
                        #'Error: No diffs 2' 
                elif ss0:
                    s0 = ss0['sys']
                    sc0 = ss0['children']                    
                    diffData = {'零件编号1':s0[1], '零件名称1':s0[2], '数量1':s0[3], '层级1':s0[0],'工程师1':'','备注1':'','零件编号2':'/', '零件名称2':'/', '数量2':'/', '层级2':'/','备注2':''}
                    diffDatas.append(diffData)
                    for c0 in sc0:
                        diffData = {'零件编号1':'/', '零件名称1':'/', '数量1':'/', '层级1':'/','工程师1':'/','备注1':'','零件编号2':'/', '零件名称2':'/', '数量2':'/', '层级2':'/','备注2':''}
                        diffData['零件编号1'] = c0[1]
                        diffData['零件名称1'] = c0[2]
                        diffData['数量1'] = c0[3]
                        diffData['层级1'] = c0[0]
                        diffData['工程师1'] = ''
                        diffData['备注1'] = ''
                        diffDatas.append(diffData)
                elif ss1:   
                    s1 = ss1['sys']
                    sc1 = ss1['children']                    
                    diffData = {'零件编号1':'/', '零件名称1':'/', '数量1':'/', '层级1':'/','工程师1':'/','备注1':'','零件编号2':s1[1], '零件名称2':s1[2], '数量2':s1[3], '层级2':s1[0],'备注2':''}
                    diffDatas.append(diffData)
                    for c1 in sc1:
                        diffData = {'零件编号1':'/', '零件名称1':'/', '数量1':'/', '层级1':'/','工程师1':'/','备注1':'','零件编号2':'/', '零件名称2':'/', '数量2':'/', '层级2':'/','备注2':''}
                        diffData['零件编号2'] = c1[1]
                        diffData['零件名称2'] = c1[2]
                        diffData['数量2'] = c1[3]
                        diffData['层级2'] = c1[0]
                        diffData['备注2'] = ''
                        diffDatas.append(diffData)
            except Exception as e:
                'Error: 出错了！'
                #e
                col1, col2 = st.columns(2)
                with col1:
                    'ss0:',ss0
                with col2:
                    'ss1:',ss1
        return diffDatas

    def SystemChildrenDiff(ss0, ss1, cols1, cols2, container):
        s0 = ss0['sys']
        s1 = ss1['sys']
        sc0 = ss0['children']
        sc1 = ss1['children']    

        r = False
        n0 = len(sc0)
        n1 = len(sc1)
        
        for i in range(0,n0):
            b0 = sc0[i]
            id0 = b0[0]            
            bIn = False
            for j in range(0, n1):                
                b1 = sc1[j]
                id1 = b1[0]
                if (id0 == id1):
                    bIn = True
                    r =  not(str(b0[1]) == str(b1[1]) and str(b0[2]) == str(b1[2]) and str(b0[3]) == str(b1[3]))
                    if not(r):
                        break
            r = r or (not bIn)
            if r:
                break 

        if not r:
            for i in range(0,n1):
                b1 = sc1[i]
                id1 = b1[0]            
                bIn = False
                for j in range(0, n0):                
                    b0 = sc0[j]
                    id0 = b0[0]
                    if (id0 == id1):
                        bIn = True
                        r =  not(str(b0[1]) == str(b1[1]) and str(b0[2]) == str(b1[2]) and str(b0[3]) == str(b1[3]))
                        if not(r):
                            break
                r = r or (not bIn)
                if r:
                    break               

        diff2ys = []
        if r:
            cols2 = container.columns(2)
            if (s0[1] == s1[1]):
                cols2[0].markdown(':blue['+BomLineStr(s0)+']')
                cols2[1].markdown(':blue['+BomLineStr(s1)+']')
            else:
                cols2[0].markdown(':red['+BomLineStr(s0)+']')
                cols2[1].markdown(':red['+BomLineStr(s1)+']')  

            dict0 = {}
            dict1 = {}
            
            d0s = []
            u0 = {'id':s0[0], 'name':s0[1], 'count':str(s0[2]), 'level':s0[3]}
            d0s.append(u0)
            dict0[s0[0]] = u0
            for d in sc0:
                u0 = {'id':d[0], 'name':d[1], 'count':str(d[2]), 'level':d[3]}
                d0s.append(u0)
                dict0[d[0]] = u0
            d0 = pd.DataFrame(d0s)

            d1s = []
            u1 = {'id':s1[0], 'name':s1[1], 'count':str(s1[2]), 'level':s1[3]}
            d1s.append(u1)
            dict1[s1[0]] = u1
            for d in sc1:
                u1 = {'id':d[0], 'name':d[1], 'count':str(d[2]), 'level':d[3]}
                d1s.append(u1)
                dict1[d[0]] = u1
            d1 = pd.DataFrame(d1s)

            td = [s0]+sc0
            td = [{'id':s0[0], 'name':s0[1], 'count':str(s0[2]), 'level':s0[3]}]
            for c0 in sc0:
                td.append({'id':c0[0], 'name':c0[1], 'count':str(c0[2]), 'level':c0[3]})
            tree0 = generate_tree(solveTreeData(td), 0)

            td = [s1]+sc1
            td = [{'id':s1[0], 'name':s1[1], 'count':str(s1[2]), 'level':s1[3]}]
            for c1 in sc1:
                td.append({'id':c1[0], 'name':c1[1], 'count':str(c1[2]), 'level':c1[3]})
            tree1 = generate_tree(solveTreeData(td), 0)

            ids0 = dict0.keys()
            ids1 = dict1.keys()                   

            diffs = compare_trees(tree0[0], tree1[0])
            if diffs:               
                rds0 = []
                rds1 = []
                rds = []
                eds = []
                nullUnit = {'id':None, 'name':None, 'count':None, 'level':None}

                for d0 in d0s:
                    id0 = d0['id']
                    u0 = dict0[id0]
                    if (id0 in ids0) and (id0 in ids1):
                        if not (id0 in eds):                            
                            u1 = dict1[id0]
                            rds0.append(dict0[id0])
                            rds1.append(dict1[id0])
                            rds.append({'id0':u0['id'],'name0':u0['name'],'count0':u0['count'],'level0':u0['level'], 'id1':u1['id'],'name1':u1['name'],'count1':u1['count'],'level1':u1['level']})
                            eds.append(id0)
                    else:
                        rds0.append(dict0[id0])                        
                        rds1.append(nullUnit)
                        rds.append({'id0':u0['id'],'name0':u0['name'],'count0':u0['count'],'level0':u0['level'], 'id1':None,'name1':None,'count1':None,'level1':None})
                for d1 in d1s:
                    id1 = d1['id']
                    u1 = dict1[id1]
                    if (id1 in ids0) and (id1 in ids1):
                        if not (id1 in eds):
                            u0 = dict0[id1]
                            rds0.append(dict0[id1])
                            rds1.append(dict1[id1])
                            rds.append({'id0':u0['id'],'name0':u0['name'],'count0':u0['count'],'level0':u0['level'], 'id1':u1['id'],'name1':u1['name'],'count1':u1['count'],'level1':u1['level']})
                            eds.append(id1)
                    else:
                        rds0.append(nullUnit)
                        rds1.append(dict1[id1])
                        rds.append( {'id0':None,'name0':None,'count0':None,'level0':None, 'id1':u1['id'],'name1':u1['name'],'count1':u1['count'],'level1':u1['level']})
                d0 = pd.DataFrame(rds0)
                d1 = pd.DataFrame(rds1)                
                color = ((d0.name != d1.name) | (d0['count'] != d1['count']) | (d0.level != d1.level)).map({True: 'background-color: yellow', False: ''})
                d0 = d0.style.apply(lambda s: color)
                d1 = d1.style.apply(lambda s: color)
                d = pd.DataFrame(rds)
                color = ((d.name0 != d.name1) | (d.count0 != d.count1) | (d.level0 != d.level1)).map({True: 'background-color: yellow', False: ''})
                d = d.style.apply(lambda s: color)
                container.dataframe(d, use_container_width=True)                          
                        
                diffDatas0 = [] #[{'零件编号':s0[0], '零件名称':s0[1], '数量':s0[2], '层级':s0[3]}]
                diffDatas1 = [] #[{'零件编号':s1[0], '零件名称':s1[1], '数量':s1[2], '层级':s1[3]}]
                for diff in diffs:
                    t1 = diff['tree1']
                    t2 = diff['tree2']
                    diffData0 = {'零件编号':None, '零件名称':None, '数量':None, '层级':None}
                    diffData1 = {'零件编号':None, '零件名称':None, '数量':None, '层级':None}
                    if t1:
                        diffData0['零件编号'] = t1['id']
                        diffData0['零件名称'] = t1['name']
                        diffData0['数量'] = t1['count']
                        diffData0['层级'] = t1['level']
                    if t2:
                        diffData1['零件编号'] = t2['id']
                        diffData1['零件名称'] = t2['name']
                        diffData1['数量'] = t2['count']
                        diffData1['层级'] = t2['level']
                    diffDatas0.append(diffData0)
                    diffDatas1.append(diffData1)
                cols2 = container.columns(2)
                cols2[0].dataframe(diffDatas0, use_container_width=True)
                cols2[1].dataframe(diffDatas1, use_container_width=True)
        else:
            if not (s0[1] == s1[1]):
                cols1[0].markdown(':red['+BomLineStr(s0)+']')
                cols1[1].markdown(':red['+BomLineStr(s1)+']')

    def GetSystemHash(sys):
        units = []
        sstr = ''
        for s in sys:
            if not (s['level'] == '2Y'):
                unit = '%s-%s-%s' % (s['id'],s['count'],s['level'])
                units.append(unit)
                sstr = ''.join(sorted(units))
        return md5(sstr.encode()).hexdigest()

    def RefreshUnitsFromEBom(uploaded_file, unit_db, bom_db, sys_db):
        unit = Query()
        df = pd.read_excel(uploaded_file, sheet_name=None)
        boms = {}
        pr = st.progress(0, '')
        c2Y = ''
        b2Y = True
        for sheetName in df.keys():
            try:
                df = pd.read_excel(uploaded_file, header=9, usecols=['零件编号','零件名称','数量','层级'], sheet_name=sheetName)
                bomArray = df.values
                systems = {}
                system = None
                n = len(bomArray)
                i = 0
                pText = '加载Ebom[%s] (%d/%d)'%(sheetName, i, n)
                for bomLine in bomArray:
                    uId = bomLine[0]
                    uName = bomLine[1]
                    uCount = bomLine[2]
                    uLevel = bomLine[3]
                    r = unit_db.get(unit.id == uId)                
                    if uLevel == '2Y':                    
                        c2Y = uId
                        if sys_db.contains(unit.id == uId):
                            #st.markdown(':red[BOM：%s 2Y系统：%s 已存在，跳过零件：%s - %s - %s~]' % (sheetName, c2Y, uId, uName, uLevel))
                            b2Y = False    
                        else:                        
                            b2Y = True
                            pId = uId
                            sys_db.insert({'id': uId, 'name': uName, 'count':uCount, 'level':uLevel, 'parent':None, 'bom':sheetName})
                            #st.markdown(':blue[新增 BOM：%s 2Y系统：%s 增加零件：%s - %s - %s~]' % (sheetName, c2Y, uId, uName, uLevel))
                    else:  
                        if b2Y:
                            sys_db.insert({'id': uId, 'name': uName, 'count':uCount, 'level':uLevel, 'parent':pId, 'bom':sheetName})
                            #st.markdown(':green[增加零件：%s - %s - %s~]' % (uId, uName, uLevel))                        
                        #else:
                            #st.markdown(':blue[BOM：%s 2Y系统：%s 已存在，跳过零件：%s - %s - %s~]' % (sheetName, c2Y, uId, uName, uLevel))
                    i += 1
                    pr.progress(float(i)/float(n), text='正在处理【%s】数据(%d/%d)'%(sheetName, i, n))
                boms[sheetName] = systems
            except Exception as e:
                st.write('出错了: 问题数据表名称: 【%s】 可能不符合标准的BOM数据表格式' % sheetName)

    def RefreshBomSystemHash():
        sys_db = TinyDB('sys_db.json')
        Unit = Query()
        r = sys_db.search(Unit.level == '2Y')
        i = 0
        n = len(r)
        pr = st.progress(0, '')
        for s in r:
            sid = s['id']
            sbom = s['bom']
            sys = sys_db.search((Unit.parent == sid) & (Unit.bom == sbom))
            sysHash = GetSystemHash(sys)
            s['hash'] = sysHash
            sys_db.upsert(s, Unit.id == sid)
            i += 1
            #st.write('Update Hash: %s  -  %s (%d / %d)' % (sid, sysHash, i, n))
            pr.progress(float(i)/float(n), text='更新2Y系统Hash值(%d/%d)'%(i, n))
        pr.progress(1.0, text='更新完成！')

    def solveTreeData(treeData):
        treeDatas = []
        lastTreeNodes = {2:treeData[0]}
        lastErrId = ''
        for data in treeData:
            levelStr = str(data['level'])
            isNode = False
            level = 0
            leafData = {}        
            if 'Y' in levelStr:
                isNode = True
                level = int(levelStr.split('Y')[0])
                if level > 2:
                    lastTreeNodes[level] = data
                    leafData = {'parent':lastTreeNodes[level-1]['id'], 'data':data}
                else:
                    leafData = {'parent':0, 'data':data}
            else:
                try:
                    level = int(levelStr)
                    leafData = {'parent':lastTreeNodes[level]['id'], 'data':data}
                except:
                    level = 2
                    eid = lastTreeNodes[level]['id']
                    leafData = {'parent':eid, 'data':data}
                    if not (lastErrId == eid):
                        lastErrId = eid
                        st.markdown(':red[严重数据错误!!!]')
                        #'leafData:', leafData
                        #'lastErrId:', lastErrId
                        #'lastTreeNodes:', lastTreeNodes
                        st.dataframe(treeData)
                
            treeDatas.append(leafData) 
        return treeDatas
    
    def solveTreeData_MBOM(treeData):
        treeDatas = []
        lastTreeNodes = {1:treeData[0]}
        lastErrId = ''
        for data in treeData:
            level = int(data['level'])
            leafData = {}
            if level > 1:
                lastTreeNodes[level] = data
                leafData = {'parent':lastTreeNodes[level-1]['id'], 'data':data}
            else:
                leafData = {'parent':0, 'data':data}
            treeDatas.append(leafData) 
        return treeDatas
                
    def generate_tree(source, parent):
        tree = []
        for item in source:
            data = item['data']
            id = data["id"]
            if item["parent"] == parent:
                item["children"] = generate_tree(source, id)
                tree.append(item)
        return tree

    def CheckErrorSystems():
        sys_db = TinyDB('sys_db.json')
        Unit = Query()
        r = sys_db.search(Unit.level == '2Y')
        i = 0
        n = len(r)
        ckds = []
        loadBar = st.progress(0, text='')
        k = 0
        for s in r:
            sid = s['id']
            sHash = s['hash']
            ss = sys_db.search((Unit.hash == sHash))
            m = len(ss)
            if not(sHash in ckds):
                if len(ss) > 1:
                    k += 1
                    st.markdown(':red[Error: 发现%d个2Y系统具有相同的结构]' % m) 
                    ckds.append(sHash)
                    cols = st.columns(m)
                    ii = 0                
                    for sys in ss:
                        r1 = sys_db.search((Unit.id == sys['id']) | (Unit.parent == sys['id']))
                        cols[ii].dataframe(r1)
                        ii += 1
            i += 1
            pText = '核查进度： %s (%d/%d)'%(sid, i, n)
            loadBar.progress(float(i)/float(n), text=pText)
        loadBar.progress(1.0, text='核对完毕，共发现%d个问题请查阅！' % k)
            
    def ShowBomDatabaseApp():
        unit_db = TinyDB('unit_db.json')
        bom_db = TinyDB('bom_db.json')
        sys_db = TinyDB('sys_db.json')
        st.title('BOM数据库')
        Unit = Query()

        with st.expander('数据库更新'):
            st.subheader('数据库更新')
            uploaded_file = st.file_uploader("通过上传Ebom清单批量更新数据库", type=["xlsx"])
            if uploaded_file is not None:
                RefreshUnitsFromEBom(uploaded_file, unit_db, bom_db, sys_db)

            if st.button('更新Hash'):
                RefreshBomSystemHash()

            #st.subheader('输入单个零部件信息')
            #cols = st.columns(4)
            #uId = cols[0].text_input('零件编号', value="")
            #uName = cols[1].text_input('零件名称', value="")
            #uCount = cols[2].number_input('数量', value=1, format='%d')
            #uLevel = cols[3].text_input('层级', value="")
            #if st.button('确定', type='primary'):
            #    u = {'id': uId, 'name': uName}
            #    if uId:                
            #        r = unit_db.get(Unit.id == uId)
            #        if r:
            #            if r['name'] == uName:
            #                st.dataframe(r)
            #                st.write('零部件已存在，且数据无差异，无需更新')
            #            else:
            #                cols = st.columns(3)
            #                cols[0].write('原零件名称：%s , 新零件名称：%s, 是否更新？' % (r['name'], uName))
            #                def on_refresh():
            #                    r = unit_db.upsert({'id': uId, 'name': uName}, Unit.id == uId)
            #                
            #                cols[1].button('更新', on_click=on_refresh, type='primary')
            #                cols[2].button('取消')  
            #        else:
            #            r = unit_db.upsert({'id': uId, 'name': uName}, Unit.id == uId)
            #            if r:  
            #                st.write(unit_db.get(Unit.id == uId))
            #    else:
            #        st.write('零件编号不可为空！')

                
        #with st.expander('零件数据查询'):
        #    st.subheader('零件数据查询')
        #    uId = st.text_input('请输入零件编号（支持模糊查询）', value="")
        #    if uId:
        #        r = unit_db.search(Unit.id.matches('.*%s.*' % uId))
        #        st.dataframe(r)
        #    uName = st.text_input('请输入零件名称（支持模糊查询）', value="")
        #    if uName:
        #        r = unit_db.search(Unit.name.matches('.*%s.*' % uName))
        #        st.dataframe(r)

        with st.expander('2Y结构查询'):
            st.subheader('2Y结构查询')
            cols1, cols2 = st.columns(2)
            uId = cols1.text_input('请输入2Y系统零件编号（支持模糊查询）1', value="")
            tree1 = None
            tree2 = None
            if uId:
                r = sys_db.search(Unit.id.matches('.*%s.*' % uId) & (Unit.level == '2Y'))
                ids = []
                for s in r:
                    ids.append(s['id'])
                sid = cols1.selectbox('选择2Y系统零件号1', ids)
                r = sys_db.search((Unit.id == sid) | (Unit.parent == sid))

                cols1.dataframe(r)
                    
                tree1 = generate_tree(solveTreeData(r), 0)
                #st.write(tree1[0])
            uId = cols2.text_input('请输入2Y系统零件编号（支持模糊查询）2', value="")
            if uId:
                r = sys_db.search(Unit.id.matches('.*%s.*' % uId) & (Unit.level == '2Y'))
                ids = []
                for s in r:
                    ids.append(s['id'])
                sid = cols2.selectbox('选择2Y系统零件号2', ids)
                r = sys_db.search((Unit.id == sid) | (Unit.parent == sid))

                cols2.dataframe(r)
                    
                tree2 = generate_tree(solveTreeData(r), 0)
                #st.write(bTreeData)
            if tree1 and tree2:
                if st.button('比较两个2Y差异'):
                    diffDatas = []
                    diffs = compare_trees(tree1[0], tree2[0])
                    if diffs:
                        for diff in diffs:
                            t1 = diff['tree1']
                            t2 = diff['tree2']
                            diffData = {'id1':None, 'name1':None, 'count1':None, 'level1':None,
                                        'id2':None, 'name2':None, 'count2':None, 'level2':None}
                            #st.write(t1)
                            #st.write(t2)
                            if t1:
                                diffData['id1'] = t1['id']
                                diffData['name1'] = t1['name']
                                diffData['count1'] = t1['count']
                                diffData['level1'] = t1['level']
                            if t2:
                                diffData['id2'] = t2['id']
                                diffData['name2'] = t2['name']
                                diffData['count2'] = t2['count']
                                diffData['level2'] = t2['level']
                            diffDatas.append(diffData)
                    if diffDatas:        
                        st.dataframe(diffDatas, use_container_width=True)
                    else:
                        st.write('无差异')

    def ShowBomDatabaseProblems():
        with st.expander('核对不同2Y号具有相同结构的问题'):
            if st.button('开始核查'):
                CheckErrorSystems()

    def showEBOMAndMBOMDiffSheetTool():
        st.header('EBOM&MBOM差异件清单生成工具')
        eboms = {}
        mbom = {}
        ebomFile = ''
        mbomFile = ''
        if 'emdiff_ebomFile' in st.session_state:
            ebomFile = st.session_state['emdiff_ebomFile']
        if 'emdiff_mbomFile' in st.session_state:
            mbomFile = st.session_state['emdiff_mbomFile']
        if 'emdiff_eboms' in st.session_state:
            eboms = st.session_state['emdiff_eboms']
        if 'emdiff_mbom' in st.session_state:
            mbom = st.session_state['emdiff_mbom']
        col1,col2 = st.columns(2)
        loadBar = st.progress(0, text='')
        uploaded_ebomFile = col1.file_uploader("上传EBOM清单", type=["xlsx"])
        uploaded_mbomFile = col2.file_uploader("上传MBOM清单", type=["xlsx"])
        if uploaded_ebomFile is not None:
            fname = uploaded_ebomFile.name       
            if not (ebomFile == fname):
                st.session_state['emdiff_ebomFile'] = fname
                df = pd.read_excel(uploaded_ebomFile, header=9)
                df['数量'] = df.apply(lambda x:getFixCountStr(x), axis=1)
                try:
                    loc1 = df.columns.get_loc('备注')
                    loc2 = df.columns.get_loc('CMAN')
                    tpNames = []
                    ks = df.keys()
                    for i in range(loc1+1, loc2):
                        tpNames.append(ks[i])
                    keys = ['零件编号','零件名称','数量','层级','负责人员'] + tpNames
                    tdf = df[keys]
                    for tpName in tpNames:
                        tpDf = tdf[tdf[tpName] == '●']
                        bomArray = tpDf.values
                        systems = {}
                        system = None
                        n = len(bomArray)
                        i = 0
                        for bomLine in bomArray:
                            sysName = bomLine[1]
                            sysLevel = bomLine[3]
                            if sysLevel == '2Y':
                                system = {'sys':bomLine , 'children':[]}
                                systems[sysName] = system
                            else:
                                if system:
                                    system['children'].append(bomLine)
                            i += 1
                            pText = '加载车型[%s] (%d/%d) 完成！'% (tpName, i, n)
                            loadBar.progress(float(i)/float(n), text='')
                        eboms[tpName] = systems                    
                except Exception as e:
                    #pText = '加载Ebom[%s]失败，格式不正确'%sheetName
                    #st.write(pText)
                    e
                    pass
                finally:
                    st.session_state['emdiff_eboms'] = eboms
            ebomFile = fname
            st.success('EBOM:%s已上传完成' % fname)

        'eboms', eboms

        if uploaded_mbomFile is not None:
            fname = uploaded_mbomFile.name            
            if not (mbomFile == fname):
                try:
                    st.session_state['emdiff_mbomFile'] = fname
                    df = pd.read_excel(uploaded_mbomFile, usecols=[0,2,3,4])
                    df.columns = ['级别','对象标识','对象描述','组件数量(CUn)']
                    tdf = df
                    bomArray = tdf.values
                    systems = {}
                    system = None
                    i = 0
                    n = len(bomArray)
                    for bomLine in bomArray:
                        sysName = bomLine[2]
                        sysLevel = bomLine[0]
                        if sysLevel == 1:
                            system = {'sys':bomLine , 'children':[]}
                            systems[sysName] = system
                        else:
                            if system:
                                system['children'].append(bomLine)
                        loadBar.progress(float(i)/float(n), text='')
                    mbom = systems
                except Exception as e:
                    #pText = 'MBOM[%s]失败，格式不正确'%sheetName
                    #st.write(pText)
                    e
                    pass
                finally:
                    st.session_state['emdiff_mbom'] = mbom
            mbomFile = fname
            st.success('MBOM: %s 已上传完成!' % fname) 

        if 0:
            'mbom', mbom


            bom0 = mbom1
            bom1 = mbom2

            if mbom1 and mbom2:
                loadBar = st.progress(0, text='')
                #cols = st.columns(2)
                #cols[0].subheader(mbomFile1)
                #cols[1].subheader(mbomFile2)
                ep = st.expander("差异件清单")
                cols = ep.columns(2)
                
                sns0 = []
                sns1 = []
                allDiffData = []

                sns0 = list(bom0.keys())
                sns1 = list(bom1.keys())           
                
                n = len(sns1)
                i = 0
                for sn in sns1:
                    ss1 = bom1[sn]
                    sid1 = ss1['sys'][1]
                    if sn in sns0:                        
                        ss0 = bom0[sn]
                        r = SystemDiff_MBOM(ss0, ss1, ep)
                        allDiffData += r                  
                    else:
                        allDiffData += SystemDiff_MBOM(None, ss1, ep)                        
                    i += 1                    
                    pText = '生成进度： (%d/%d)'%(i, n)
                    loadBar.progress(float(i)/float(n), text=pText)

                n = len(sns0)
                i = 0
                for sn in sns0:
                    if sn in sns1:
                        continue
                    ss0 = bom0[sn]                  
                    r = SystemDiff_MBOM(ss0, None, ep)
                    allDiffData += r
                    i += 1                    
                    pText = '生成进度： (%d/%d)'%(i, n)
                    loadBar.progress(float(i)/float(n), text=pText)
                loadBar.progress(1.0, text='MBOM差异件清单生成完成!')

                if allDiffData:
                    df = pd.DataFrame(allDiffData)
                    color = ((df['层级1'] == 1) | (df['层级2'] == 1)).map({True: 'background-color: #EEEE99', False: ''})
                    cdf = df.style.apply(lambda s: color)
                    df = df.set_index('零件编号1')
                    ep.dataframe(cdf, use_container_width=True)
                    import io
                    buffer = io.BytesIO()
                    ef1 = mbomFile1.lower().split('.xlsx')[0]
                    ef2 = mbomFile2.lower().split('.xlsx')[0]
                    with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
                        sheet_new = 'Sheet1'
                        cdf.to_excel(writer, sheet_name=sheet_new, index=False)                  
                        writer.close()
                        fileName= '%s_%s_差异件清单.xlsx' % (ef1, ef2)
                        ep.write(fileName)
                        ep.download_button(
                            label="下载差异件清单",
                            data=buffer,
                            file_name=fileName,
                            mime='application/vnd.ms-excel'
                            )  
                else:
                    ep.write('所选两个MBOM无差异') 
            else:
                '请加载要对比的MBOM清单 '


    def showMBOMDiffSheetTool():
        st.header('MBOM差异件清单生成工具')
        mbom1 = {}
        mbom2 = {}
        mbomFile1 = ''
        mbomFile2 = ''
        if 'mbomFile1' in st.session_state:
            mbomFile1 = st.session_state['mbomFile1']
        if 'mbomFile2' in st.session_state:
            mbomFile2 = st.session_state['mbomFile2']
        if 'mbom1' in st.session_state:
            mbom1 = st.session_state['mbom1']
        if 'mbom2' in st.session_state:
            mbom2 = st.session_state['mbom2']
        col1,col2 = st.columns(2)
        loadBar = st.progress(0, text='')
        uploaded_file1 = col1.file_uploader("上传MBOM清单1", type=["xlsx"])
        uploaded_file2 = col2.file_uploader("上传MBOM清单2", type=["xlsx"])
        if uploaded_file1 is not None:
            fname = uploaded_file1.name       
            if not (mbomFile1 == fname):
                try:
                    st.session_state['mbomFile1'] = fname
                    df = pd.read_excel(uploaded_file1, usecols=[0,2,3,4])
                    df.columns = ['级别','对象标识','对象描述','组件数量(CUn)']
                    tdf = df
                    bomArray = tdf.values
                    systems = {}
                    system = None
                    i = 0
                    n = len(bomArray)
                    for bomLine in bomArray:
                        sysName = bomLine[2]
                        sysLevel = bomLine[0]
                        if sysLevel == 1:
                            system = {'sys':bomLine , 'children':[]}
                            systems[sysName] = system
                        else:
                            if system:
                                system['children'].append(bomLine)
                        loadBar.progress(float(i)/float(n), text='')
                    mbom1 = systems
                except Exception as e:
                    #pText = '加载MBOM[%s]失败，格式不正确'%sheetName
                    #st.write(pText)
                    e
                    pass
                finally:
                    st.session_state['mbom1'] = mbom1
            mbomFile1 = fname
            st.success('MBOM1 已上传完成')

        if uploaded_file2 is not None:
            fname = uploaded_file2.name            
            if not (mbomFile2 == fname):
                try:
                    st.session_state['mbomFile2'] = fname
                    df = pd.read_excel(uploaded_file2, usecols=[0,2,3,4])
                    df.columns = ['级别','对象标识','对象描述','组件数量(CUn)']
                    tdf = df
                    bomArray = tdf.values
                    systems = {}
                    system = None
                    i = 0
                    n = len(bomArray)
                    for bomLine in bomArray:
                        sysName = bomLine[2]
                        sysLevel = bomLine[0]
                        if sysLevel == 1:
                            system = {'sys':bomLine , 'children':[]}
                            systems[sysName] = system
                        else:
                            if system:
                                system['children'].append(bomLine)
                        loadBar.progress(float(i)/float(n), text='')
                    mbom2 = systems
                except Exception as e:
                    #pText = 'MBOM[%s]失败，格式不正确'%sheetName
                    #st.write(pText)
                    e
                    pass
                finally:
                    st.session_state['mbom2'] = mbom2
            mbomFile2 = fname
            st.success('MBOM2 已上传完成') 

        bom0 = mbom1
        bom1 = mbom2

        if mbom1 and mbom2:
            loadBar = st.progress(0, text='')
            #cols = st.columns(2)
            #cols[0].subheader(mbomFile1)
            #cols[1].subheader(mbomFile2)
            ep = st.expander("差异件清单")
            cols = ep.columns(2)
            
            sns0 = []
            sns1 = []
            allDiffData = []

            sns0 = list(bom0.keys())
            sns1 = list(bom1.keys())           
            
            n = len(sns1)
            i = 0
            for sn in sns1:
                ss1 = bom1[sn]
                sid1 = ss1['sys'][1]
                if sn in sns0:                        
                    ss0 = bom0[sn]
                    r = SystemDiff_MBOM(ss0, ss1, ep)
                    allDiffData += r                  
                else:
                    allDiffData += SystemDiff_MBOM(None, ss1, ep)                        
                i += 1                    
                pText = '生成进度： (%d/%d)'%(i, n)
                loadBar.progress(float(i)/float(n), text=pText)

            n = len(sns0)
            i = 0
            for sn in sns0:
                if sn in sns1:
                    continue
                ss0 = bom0[sn]                  
                r = SystemDiff_MBOM(ss0, None, ep)
                allDiffData += r
                i += 1                    
                pText = '生成进度： (%d/%d)'%(i, n)
                loadBar.progress(float(i)/float(n), text=pText)
            loadBar.progress(1.0, text='MBOM差异件清单生成完成!')

            if allDiffData:
                df = pd.DataFrame(allDiffData)
                color = ((df['层级1'] == 1) | (df['层级2'] == 1)).map({True: 'background-color: #EEEE99', False: ''})
                cdf = df.style.apply(lambda s: color)
                df = df.set_index('零件编号1')
                ep.dataframe(cdf, use_container_width=True)
                import io
                buffer = io.BytesIO()
                ef1 = mbomFile1.lower().split('.xlsx')[0]
                ef2 = mbomFile2.lower().split('.xlsx')[0]
                with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
                    sheet_new = 'Sheet1'
                    cdf.to_excel(writer, sheet_name=sheet_new, index=False)                  
                    writer.close()
                    fileName= '%s_%s_差异件清单.xlsx' % (ef1, ef2)
                    ep.write(fileName)
                    ep.download_button(
                        label="下载差异件清单",
                        data=buffer,
                        file_name=fileName,
                        mime='application/vnd.ms-excel'
                        )  
            else:
                ep.write('所选两个MBOM无差异') 
        else:
            '请加载要对比的MBOM清单 '

    def getFixCountStr(line):
        cnt = line['数量']
        try:
            cntInt = int(cnt)
            if cnt == cntInt:
                cnt = str(cntInt)
            else:
                cnt = str(cnt)
        except Exception as e:
             cnt = str(cnt)
        finally:
            return cnt

    def showDiffSheetTool():
        st.header('BOM差异件清单生成工具') 

        eboms1 = {}
        eboms2 = {}
        ebomFile1 = ''
        ebomFile2 = ''

        if 'ebomFile1' in st.session_state:
            ebomFile1 = st.session_state['ebomFile1']
        if 'ebomFile2' in st.session_state:
            ebomFile2 = st.session_state['ebomFile2']
        if 'eboms1' in st.session_state:
            eboms1 = st.session_state['eboms1']
        if 'eboms2' in st.session_state:
            eboms2 = st.session_state['eboms2']

        col1,col2 = st.columns(2)

        loadBar = st.progress(0, text='')

        uploaded_file1 = col1.file_uploader("上传EBOM清单1", type=["xlsx"])
        uploaded_file2 = col2.file_uploader("上传EBOM清单2", type=["xlsx"])
        if uploaded_file1 is not None:
            fname = uploaded_file1.name            
            if not (ebomFile1 == fname):
                st.session_state['ebomFile1'] = fname
                df = pd.read_excel(uploaded_file1, header=9)
                df['数量'] = df.apply(lambda x:getFixCountStr(x), axis=1)
                try:
                    loc1 = df.columns.get_loc('备注')
                    loc2 = df.columns.get_loc('CMAN')
                    tpNames = []
                    ks = df.keys()
                    for i in range(loc1+1, loc2):
                        tpNames.append(ks[i])
                    keys = ['零件编号','零件名称','数量','层级','负责人员'] + tpNames
                    tdf = df[keys]
                    for tpName in tpNames:
                        #tpDf = tdf.query('%s == "●"' % tpName)
                        tpDf = tdf[tdf[tpName] == '●']
                        bomArray = tpDf.values
                        systems = {}
                        system = None
                        n = len(bomArray)
                        i = 0
                        for bomLine in bomArray:
                            sysName = bomLine[1]
                            sysLevel = bomLine[3]
                            if sysLevel == '2Y':
                                system = {'sys':bomLine , 'children':[]}
                                systems[sysName] = system
                            else:
                                if system:
                                    system['children'].append(bomLine)
                            i += 1
                            pText = '加载车型[%s] (%d/%d) 完成！'% (tpName, i, n)
                            loadBar.progress(float(i)/float(n), text='')
                        eboms1[tpName] = systems
                except Exception as e:
                    #pText = '加载Ebom[%s]失败，格式不正确'%sheetName
                    #st.write(pText)
                    e
                    pass
                finally:
                    st.session_state['eboms1'] = eboms1
            ebomFile1 = fname
            st.success('EBOM1:%s已上传完成' % fname)
        if uploaded_file2 is not None:
            fname = uploaded_file2.name
            if not (ebomFile2 == fname):
                st.session_state['ebomFile2'] = fname
                df = pd.read_excel(uploaded_file2, header=9) 
                df
                df['数量'] = df.apply(lambda x:getFixCountStr(x), axis=1)
                
                try:
                    loc1 = df.columns.get_loc('备注')
                    loc2 = df.columns.get_loc('CMAN')
                    tpNames = []
                    ks = df.keys()
                    for i in range(loc1+1, loc2):
                        tpNames.append(ks[i])
                    keys = ['零件编号','零件名称','数量','层级','负责人员'] + tpNames                    
                    tdf = df[keys]                    
                    for tpName in tpNames:
                        #tpDf = tdf.query('%s == "●"' % tpName)
                        tpDf = tdf[tdf[tpName] == '●']
                        bomArray = tpDf.values
                        systems = {}
                        system = None
                        n = len(bomArray)
                        i = 0
                        for bomLine in bomArray:
                            sysName = bomLine[1]
                            sysLevel = bomLine[3]
                            if sysLevel == '2Y':
                                system = {'sys':bomLine , 'children':[]}
                                systems[sysName] = system
                            else:
                                if system:
                                    system['children'].append(bomLine)
                            i += 1
                            pText = '加载车型[%s] (%d/%d) 完成！'% (tpName, i, n)
                            loadBar.progress(float(i)/float(n), text='')
                        eboms2[tpName] = systems
                except Exception as e:
                    #pText = '加载Ebom[%s]失败，格式不正确'%sheetName
                    #st.write(pText)
                    e
                    pass
                finally:
                    st.session_state['eboms2'] = eboms2
            ebomFile2 = fname
            st.success('EBOM2:%s已上传完成' % fname)      
        bomKeys1 = list(eboms1.keys())
        bomKeys2 = list(eboms2.keys())        
        if bomKeys1 and bomKeys2:        
            st.subheader('通过两个ebom生成差异件清单')
            cols = st.columns(2)            
            bk1 = cols[0].selectbox('选择第一个ebom中的版型：',bomKeys1)
            bk2 = cols[1].selectbox('选择第二个ebom中的版型：',bomKeys2)
            st.write('BOM清单比对结果：')
            loadBar = st.progress(0, text='')
            cols = st.columns(2)
            cols[0].header(bk1)
            cols[1].header(bk2)
            ep = st.expander("差异件清单")
            cols = ep.columns(2)
            bom0 = eboms1[bk1]
            bom1 = eboms2[bk2]
            sns0 = list(bom0.keys())
            sns1 = list(bom1.keys())
            
            allDiffData = []

            n = len(sns1)
            i = 0
            for sn in sns1:
                ss1 = bom1[sn]
                sid1 = ss1['sys'][0]                    
                if sn in sns0:                        
                    ss0 = bom0[sn]
                    r = SystemDiff(ss0, ss1, ep)
                    allDiffData += r                  
                else:
                    allDiffData += SystemDiff(None, ss1, ep)                        
                i += 1                    
                pText = '生成进度： (%d/%d)'%(i, n)
                loadBar.progress(float(i)/float(n), text=pText)

            n = len(sns0)
            i = 0
            for sn in sns0:
                if sn in sns1:
                    continue
                ss0 = bom0[sn]                  
                r = SystemDiff(ss0, None, ep)
                allDiffData += r
                i += 1                    
                pText = '生成进度： (%d/%d)'%(i, n)
                loadBar.progress(float(i)/float(n), text=pText)
            loadBar.progress(1.0, text='【:blue[%s]】与【:blue[%s]】差异件清单生成完成!' % (bk1, bk2))
            if allDiffData:
                df = pd.DataFrame(allDiffData)
                color = ((df['层级1'] == '2Y') | (df['层级2'] == '2Y')).map({True: 'background-color: #EEEE99', False: ''})
                cdf = df.style.apply(lambda s: color)
                df = df.set_index('零件编号1')
                ep.dataframe(cdf, use_container_width=True)
                import io
                buffer = io.BytesIO()
                ef1 = ebomFile1.split('.xlsx')[0]
                ef2 = ebomFile2.split('.xlsx')[0]
                with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
                    #workbook = writer.book
                    #formatCenter = workbook.add_format({'align':'center','font_color': '#0000FF'})
                    sheet_new = 'Sheet1'
                    cdf.to_excel(writer, sheet_name=sheet_new, index=False)
                    #worksheet=writer.sheets[sheet_new]
                    #worksheet.set_column('C:E', 7, formatCenter)
                    #worksheet.set_column('I:J', 7, formatCenter)                    
                    writer.close()
                    fileName= '%s(%s)_%s(%s)_差异件清单.xlsx' % (ef1, bk1, ef2, bk2)
                    ep.write(fileName)
                    ep.download_button(
                        label="下载差异件清单",
                        data=buffer,
                        file_name=fileName,
                        mime='application/vnd.ms-excel'
                        )  
            else:
                ep.write('所选两个配置EBOM无差异')                       

    def ShowBomDiffSheets():
        boms = {}
        bomFile = ''
        if 'bomFile' in st.session_state:
            bomFile = st.session_state['bomFile']
        if 'bomData' in st.session_state:
            boms = st.session_state['bomData']
        
        st.header('BOM表单之间数据进行核对')

        uploaded_file = st.file_uploader("上传需要核对的EBOM清单", type=["xlsx"])
        if uploaded_file is not None:
            fname = uploaded_file.name
            if not (bomFile == fname):
                st.session_state['bomFile'] = fname
                df = pd.read_excel(uploaded_file, sheet_name=None)
                boms = {}
                loadBar = st.progress(0, text='')
                for sheetName in df.keys():
                    try:
                        df = pd.read_excel(uploaded_file, header=9, usecols=['零件编号','零件名称','数量','层级'], sheet_name=sheetName)
                        bomArray = df.values
                        systems = {}
                        system = None
                        n = len(bomArray)
                        i = 0
                        pText = '加载Ebom[%s] (%d/%d)'%(sheetName, i, n)
                        for bomLine in bomArray:
                            sysId = bomLine[0]
                            sysLevel = bomLine[3]
                            if sysLevel == '2Y':
                                system = {'sys':bomLine , 'children':[]}
                                systems[sysId] = system
                            else:
                                if system:
                                    system['children'].append(bomLine)
                            i += 1
                            pText = '加载Ebom[%s] (%d/%d) 完成！'%(sheetName, i, n)
                            loadBar.progress(float(i)/float(n), text='')
                        st.write(pText)
                        boms[sheetName] = systems
                    except:
                        #pText = '加载Ebom[%s]失败，格式不正确'%sheetName
                        #st.write(pText)
                        pass
                    finally:
                        st.session_state['bomData'] = boms
                #st.write(boms)
            bomKeys = list(boms.keys())
            st.subheader('通过对比不同Bom清单核查问题')
            cols = st.columns(2)            
            bk1 = cols[0].selectbox('选择第一个BOM：',bomKeys)
            bk2 = cols[1].selectbox('选择第二个BOM：',bomKeys)
            if bk1 == bk2:
                st.write('请选择两个不同的BOM进行比对！')
            else:
                st.write('BOM清单比对结果：')
                loadBar = st.progress(0, text='')
                cols = st.columns(2)
                cols[0].header(bk1)
                cols[1].header(bk2)
                ep1 = st.expander("2Y")
                ep2 = st.expander("2Y子件")
                cols1 = ep1.columns(2)
                cols2 = ep2.columns(2)
                bom0 = boms[bk1]
                bom1 = boms[bk2]
                sns0 = list(bom0.keys())
                sns1 = list(bom1.keys())
                n = len(sns1)
                i = 0
                for sn in sns1:
                    if sn in sns0:
                        ss0 = bom0[sn]
                        ss1 = bom1[sn]              
                        SystemChildrenDiff(ss0, ss1, cols1, cols2, ep2)  
                        #    cols[0].write(ss0)
                        #    cols[1].write(ss1)
                    #else:
                    #    st.write(sn + '不在' + bomKeys[0]+'Bom清单中')
                    i += 1
                    pText = '核对进度： (%d/%d)'%(i, n)
                    loadBar.progress(float(i)/float(n), text=pText)
                loadBar.progress(1.0, text='Bom清单【:blue[%s]】与【:blue[%s]】比对结束!' % (bk1, bk2))

    def showBOMDiffTool():
        st.header('核对EBOM与差异件清单中的问题')

        df_ebom = None
        df_diff = None
        checkEBomFile = ''
        checkDiffFile = ''
                 
        if 'checkEBomFile' in st.session_state:
            checkEBomFile = st.session_state['checkEBomFile']
        if 'checkDiffFile' in st.session_state:
            checkDiffFile = st.session_state['checkDiffFile']
        if 'checkEBomDF' in st.session_state:
            df_ebom = st.session_state['checkEBomDF']
        if 'checkDiffDF' in st.session_state:
            df_diff = st.session_state['checkDiffDF']

        uploaded_ebom = st.file_uploader("上传进行核对的EBOM清单", type=["xlsx"])
        if uploaded_ebom is not None:
            fname = uploaded_ebom.name
            if not (checkEBomFile == fname):
                st.write('原EBOM文件: %s 新上传EBOM文件: %s' % (checkEBomFile, fname))
                st.session_state['checkEBomFile'] = fname
                df_ebom = pd.read_excel(uploaded_ebom, header=9)
            else:
                st.write('EBOM文件: %s 已上传' % fname)
        else:
            if 'checkEBomFile' in st.session_state:
                del st.session_state['checkEBomFile']
                
        uploaded_diff = st.file_uploader("上传进行核对的差异件清单", type=["xlsx"])
        if uploaded_diff is not None:
            fname = uploaded_diff.name
            if not (checkDiffFile == fname):
                st.write('原EBOM文件: %s 新上传EBOM文件: %s' % (checkDiffFile, fname))
                st.session_state['checkDiffFile'] = fname
            else:
                st.write('EBOM文件: %s 已上传' % fname)

    def showECTTool():
        st.header('工程配置工具')  

        df_ect = pd.DataFrame()
        ectFile = ''
        if 'ectFile' in st.session_state:
            ectFile = st.session_state['ectFile']
        if 'df_ect' in st.session_state:
            df_ect = st.session_state['df_ect']

        uploaded_file = st.file_uploader("上传工程配置表", type=["xlsx"])
        if uploaded_file is not None:
            fname = uploaded_file.name
            if ectFile == fname:
                st.write('工程配置表文件: %s 已上传' % fname)                
            else:
                st.session_state['ectFile'] = fname
                df_ect = pd.read_excel(uploaded_file, header=2, usecols=[0,2,6,7,8,9,10,11])
                st.session_state['df_ect'] = df_ect
        if df_ect.empty:
            '工程配置数据为空，请上传数据'
        else:
            df_ect
            tdf = df_ect.groupby('专业部门',as_index=False).count()['专业部门']
            ops = list(tdf.values)
            dept = st.selectbox('专业部门', ops)
            #q = 'GPC == %s and FND == "%s"' % (gpc, fnd)
            q = '专业部门 == "%s"' % dept
            df_dept = df_ect.query(q)
            tdf = df_dept.groupby('分类',as_index=False).count()['分类']
            ops = list(tdf.values)
            cls = st.selectbox('分类', ops)
            q = '分类 == "%s"' % cls
            df_cls = df_dept.query(q)
            df_cls
            tdf = df_cls.groupby(['特征族代码','特征族描述（中文）'],as_index=False).count()['特征族描述（中文）']
            ops = list(tdf.values)

    def countListS(l):
        i = 0
        for v in l:
            if (v == 'S') or (v == 's'):
                i += 1
        return i
    
    def checkLVVS(lvvs):
        scnt = len(lvvs[0][1])
        lcnt = len(lvvs)
        dups = []
        for i in range(0, scnt):
            vs = []
            for j in range(0, lcnt):
                lvv = lvvs[j]
                lvid = lvv[0]
                lvvv = lvv[1]
                v = lvvv[i]
                if (v == 'S') or (v == 's'):
                    vs.append((lvid, i))
            if len(vs) > 1:
                dups.append(vs)
        return dups
    
    def getECV(df, skey):
        cols = st.columns(5)
        tdf = df.groupby('专业部门',as_index=False).count()['专业部门']
        ops = list(tdf.values)
        dept = cols[0].selectbox('专业部门', ops, key=skey+'1')
        #q = 'GPC == %s and FND == "%s"' % (gpc, fnd)
        q = '专业部门 == "%s"' % dept
        df_dept = df.query(q)
        #df_dept 

        tdf = df_dept.groupby('分类',as_index=False).count()['分类']
        ops = list(tdf.values)
        cls = cols[1].selectbox('分类', ops, key=skey+'2')
        q = '分类 == "%s"' % cls
        df_cls = df_dept.query(q)
        #df_cls
        
        tdf = df_cls.groupby(['特征组代码','特征组描述'],as_index=False).count()['特征组描述']       
        ops = list(tdf.values)
        ccs = cols[2].selectbox('特征组', ops, key=skey+'3')
        q = '特征组描述 == "%s"' % ccs
        try:
            ccs = float(ccs)
            q = '特征组描述 == %s' % ccs 
        except:
            pass        
        df_ccs = df_cls.query(q)
        #df_ccs

        tdf = df_ccs.groupby(['特征族代码','特征族描述'],as_index=False).count()['特征族描述']
        ops = list(tdf.values)
        ccv = cols[3].selectbox('特征族', ops, key=skey+'4')
        q = '特征族描述 == "%s"' % ccv
        try:
            ccv = float(ccv)
            q = '特征族描述 == %s' % ccv
        except:
            pass
        df_ccv = df_ccs.query(q)
        #df_ccv

        tdf = df_ccv.groupby(['特征值代码','特征值描述'],as_index=False).count()['特征值描述']
        ops = list(tdf.values)
        cv = cols[4].selectbox('特征值', ops, key=skey+'5')
        q = '特征值描述 == "%s"' % cv
        try:
            cv = float(cv)
            q = '特征值描述 == %s' % cv
        except:
            pass
        df_cv = df_ccv.query(q)
        
        return df_cv, cv
    
    def getECV2(df, skey): 
        tdf = df.groupby(['特征值代码','特征值描述'],as_index=False).count()['特征值描述']
        ops = list(tdf.values)
        cv = st.selectbox('特征值', ops, key=skey)
        q = '特征值描述 == "%s"' % cv
        try:
            cv = float(cv)
            q = '特征值描述 == %s' % cv
        except:
            pass
        df_cv = df.query(q)
        
        return df_cv, cv
    
    def showLOUPointTool():
        st.header('LOU打点工具')  
        df_ectPoint = pd.DataFrame()
        ectPointFile = ''
        ectPointDict = {} #配置打点信息字典
        if 'ectPointFile' in st.session_state:
            ectPointFile = st.session_state['ectPointFile']
        if 'ectPointDict' in st.session_state:
            ectPointDict = st.session_state['ectPointDict']
        if 'df_ectPoint' in st.session_state:
            df_ectPoint = st.session_state['df_ectPoint']

        uploaded_file_ect = st.file_uploader("上传工程配置表", type=["xlsx"])
        if uploaded_file_ect is not None:
            fname = uploaded_file_ect.name
            if not (ectPointFile == fname):
                st.session_state['ectPointFile'] = fname
                ectPointDict = {}                
                df = pd.read_excel(uploaded_file_ect, header=2)
                cc = len(list(df.columns))
                rows = df.values                
                for row in rows:
                    k = str(row[10])
                    v = row[12:cc]
                    ectPointDict[k] = v
                st.session_state['ectPointDict'] = ectPointDict
                df_ectPoint = df
                st.session_state['df_ectPoint'] = df_ectPoint
            st.success('工程配置表文件: %s 已上传' % fname)

        if df_ectPoint.empty:
            st.warning('请上传工程配置表')
        else:
            ectVsDict = {}
            if 'ectVsDict' in st.session_state:
                ectVsDict = st.session_state['ectVsDict']
            #ectPointDict 
            #df_ectPoint
            cc = len(list(df_ectPoint.columns))
            #rows = df.values                
            #for row in rows:
            #    k = str(row[10])
            #    v = row[12:cc] 
            pStart = 12   
            pCount = cc - pStart     

            with st.expander('特征值'):
                vn = st.number_input('特征值数量', 1, 5, 1) 
                ecvs = []
                cvs = []
                rows = []
                for i in range(0, vn):
                    #ecv, cv = getECV(df_ectPoint, 'ecv_%s' % i)  
                    ecv, cv = getECV2(df_ectPoint, 'ecv_%s' % i)           
                    ecvs.append(ecv)
                    cvs.append(str(cv))
                    row = list(ecv.values[0])[pStart:cc]
                    rows.append(row)

                n = len(rows[0])            
                row = []
                op = ''            
                if st.button('OR'):
                    row = ['-'] * pCount
                    op = '|'
                    for i in range(0, n):
                        b = False
                        for j in range(0, vn):
                            v = str(rows[j][i])
                            b = b or (v.upper() == 'S')
                        if b:
                            row[i] = 'S'  
                if st.button('AND'):
                    row = ['-'] * pCount
                    op = '&'
                    for i in range(0, n):
                        b = True
                        for j in range(0, vn):
                            v = str(rows[j][i])
                            b = b and (v.upper() == 'S')
                        if b:
                            row[i] = 'S'         
                st.dataframe(rows)
                if row:
                    vk = ('%s' % op).join(cvs)
                    '特征组合值: %s' % vk
                    st.dataframe([row])
                    ectVsDict[vk] = row
                    st.session_state['ectVsDict'] = ectVsDict
            with st.expander('组合特征值'):
                if ectVsDict:
                    #ectVsDict
                    c_vn = st.number_input('组合特征值数量', 2, 5, 2) 
                    c_cvs = []
                    c_rows = []
                    cvKeys = ectVsDict.keys()
                    for i in range(0, c_vn):
                        ccv = st.selectbox('组合特征值', cvKeys, key='ccv_%s' % i)
                        crow = ectVsDict[ccv]
                        c_cvs.append('(%s)' % ccv)                    
                        c_rows.append(crow)
                    #c_rows
                    n = len(c_rows[0])            
                    row = []
                    op = ''            
                    if st.button('OR', key='ccop_or'):
                        row = ['-'] * pCount
                        op = '|'
                        for i in range(0, n):
                            b = False
                            for j in range(0, c_vn):
                                v = str(c_rows[j][i])
                                b = b or (v.upper() == 'S')
                            if b:
                                row[i] = 'S'  
                    if st.button('AND', key='ccop_and'):
                        row = ['-'] * pCount
                        op = '&'
                        for i in range(0, n):
                            b = True
                            for j in range(0, c_vn):
                                v = str(c_rows[j][i])
                                b = b and (v.upper() == 'S')
                            if b:
                                row[i] = 'S'         
                    st.dataframe(c_rows)
                    if row:
                        vk = ('%s' % op).join(c_cvs)
                        'lou用法: %s' % vk
                        st.dataframe([row])
                        #ectVsDict[vk] = row
                        #st.session_state['ectVsDict'] = ectVsDict
                        #ectVsDict
                
    def showLOUCheckTool():
        st.header('配置打点信息一致性核查')    
        ccFile = st.file_uploader("上传解算文件", type=["xlsx"])
        headRow = 0
        rs = []
        if ccFile is not None:
            df1 = pd.read_excel(ccFile, header=headRow)
            #df1
            lous1 = list(df1['变更后LOU用法'].values)
            #'lous1, n', lous1, len(lous1)
            cols1 = list(df1.columns)
            colLen1 = len(cols1)
            #'cols1', cols1
            pNames1 = cols1[4:]
            louFile = st.file_uploader("上传LOU文件", type=["xlsx"])
            rn1 = len(df1)
            vs1 = list(df1.values)
            values1 = []
            for v in vs1:
                values1.append(v[4:])
            #'values1', values1
            if louFile is not None:
                df2 = pd.read_excel(louFile, header=headRow) 
                #df2

                cols2 = list(df2.columns)
                #'cols2', cols2
                sIdx = 0
                eIdx = 0
                for i in range(len(cols2)):
                    col2 = cols2[i]
                    if col2 == 'LOU用法':
                        sIdx = i + 2
                    elif col2 == 'EWO变更单号':
                        eIdx = i - 1
                #'sIdx, eIdx', sIdx, eIdx+1

                vs2 = list(df2.values)[6:]
                values2 = []
                for v in vs2:
                    values2.append(v[sIdx:eIdx+1])
                #'values2', values2
                
                pCols2 = cols2[sIdx:eIdx+1]

                lous2 = list(df2['LOU用法'].values)
                lous2 = lous2[6:]
                

                #pNames2 = {}
                #i = 0
                #for i in range(len(pCols2)):
                #    pCol2 = pCols2[i]
                #   pds2 = list(df2[pCol2].values)
                #    pName2 = pds2[2]
                #    ps = pName2.split('-')[1].split('内饰')
                #    pNameX = ps[0] + '内饰本色' + ps[1]
                #    i1 = -1
                #    for j in range(len(pNames1)):
                #       pName1 = pNames1[j]
                #        if pNameX in pName1:
                #            i1 = j
                #            break
                #    pNames2[pName2] = [pNameX, i, i1]
                #'pNames', pNames1, pNames2
                    
                pNames2 = []
                pNames2X = []
                i = 0
                for i in range(len(pCols2)):
                    pCol2 = pCols2[i]
                    pds2 = list(df2[pCol2].values)
                    pName2 = pds2[2]
                    ps = pName2.split('-')[1].split('内饰')
                    pName2X = ps[0] + '内饰本色' + ps[1]
                    pNames2.append(pName2)
                    pNames2X.append(pName2X)
                #'pNames', pNames2, pNames2X

                pNamesDict = {}
                for i in range(len(pNames1)):
                    pName1 = pNames1[i]
                    k = -1
                    for j in range(len(pNames2X)):
                        pName2X = pNames2X[j]
                        if pName2X in pName1:
                            k = j
                            break
                    pNamesDict[i] = k
                #'pNamesDict', pNamesDict


                rows = []
                for i in range(rn1):
                    row = {}
                    row1 = values1[i]
                    row2 = values2[i]
                    lou1 = lous1[i]
                    lou2 = lous2[i]
                    #'row1, row2', row1, row2
                    #'pNames1, pNames2', pNames1, pNames2                                            
                    #row = {'变更后LOU用法':lou1, 'LOU用法':lou2, '工程师工号':userId, '工程师姓名':userName}
                    row = {'变更后LOU用法':lou1, 'LOU用法':lou2}
                    for j in range(len(pNames1)):
                        pIdx2 = pNamesDict[j]
                        pName1 = pNames1[j]
                        pName2 = '无_%s' % j
                        v1 = row1[j]
                        v2 = '/'
                        ck = '0'
                        if pIdx2 >= 0:   
                            pName2 = pNames2[pIdx2]
                            v2 = row2[pIdx2]
                            v1 = str(v1).lower().replace(' ','')
                            v2 = str(v2).lower().replace(' ','')                            
                            if v1 == v2:
                                ck = '1'
                        row[pName1] = v1
                        row[pName2] = v2
                        row['ck_%s' % j] = ck
                    rows.append(row)
                df = pd.DataFrame(rows)
                st.dataframe(df)

                def checkFun(value):
                    v = str(value)
                    if v == '0':
                        return 'background-color: red'
                    elif v == '1':
                        return 'background-color: green'
                    else:
                        return ''
                df = df.style.applymap(checkFun)

                import io
                buffer = io.BytesIO()
                ef1 = ccFile.name.lower().split('.xlsx')[0]
                ef2 = louFile.name.lower().split('.xlsx')[0]
                with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
                    sheet_new = 'Sheet1'
                    df.to_excel(writer, sheet_name=sheet_new, index=False)                  
                    writer.close()
                    fileName= '%s_%s_校核.xlsx' % (ef1, ef2)
                    st.write(fileName)
                    st.download_button(
                        label="下载校核文件",
                        data=buffer,
                        file_name=fileName,
                        mime='application/vnd.ms-excel'
                        )

            
           
    def showLOUTool():
        st.header('LOU核查工具')
        df_sys = pd.DataFrame()
        df_ect = pd.DataFrame()
        sysFile = ''
        ectFile = ''
        ectDict = {} #配置打点信息字典
        if 'sysFile' in st.session_state:
            sysFile = st.session_state['sysFile']
        if 'df_sys' in st.session_state:
            df_sys = st.session_state['df_sys']
        if 'ectFile' in st.session_state:
            ectFile = st.session_state['ectFile']
        if 'ectDict' in st.session_state:
            ectDict = st.session_state['ectDict']
        if 'df_ect' in st.session_state:
            df_ect = st.session_state['df_ect']
        

        col1, col2 = st.columns(2)

        uploaded_file_sys = col1.file_uploader("上传CPAC&FND查询表", type=["xlsx"])
        if uploaded_file_sys is not None:
            fname = uploaded_file_sys.name
            if not (sysFile == fname):
                st.session_state['sysFile'] = fname
                df = pd.read_excel(uploaded_file_sys, header=2) 
                columns = ['归口专业','CPAC','CPAC_CN','CPAC_EN','FND','FND_CN','FND_EN','GPC','关键件/重要件','是否紧固件','法规件名称',
                        '品类代码','品类描述','平台关键总成','本色件标识','是否精确追溯件','度量单位','是否法规件','备注']
                df.columns = columns
                lines = df.values
                lastCPAC = None
                for line in lines:
                    cpac = line[1]
                    if pd.isna(cpac):
                        line[1:4] = lastCPAC                  
                    else:
                        lastCPAC = line[1:4]   
                df_sys = pd.DataFrame(lines)
                df_sys.columns = columns
                values = {'FND_EN': '/'}
                df_sys.fillna(value=values,inplace=True)
                st.session_state['df_sys'] = df_sys
            st.success('产品架构查询文件: %s 已上传' % fname)

        uploaded_file_ect = col2.file_uploader("上传工程配置表", type=["xlsx"])
        if uploaded_file_ect is not None:
            fname = uploaded_file_ect.name
            if not (ectFile == fname):
                st.session_state['ectFile'] = fname
                ectDict = {}                
                df = pd.read_excel(uploaded_file_ect, header=2)
                cc = len(list(df.columns))
                rows = df.values                
                for row in rows:
                    k = str(row[10])
                    v = row[12:cc]
                    ectDict[k] = v
                st.session_state['ectDict'] = ectDict
                df = pd.read_excel(uploaded_file_ect, header=2, usecols=[10])                 
                #columns = ['归口专业','CPAC','CPAC_CN','CPAC_EN','FND','FND_CN','FND_EN','GPC','关键件/重要件','是否紧固件','法规件名称',
                #        '品类代码','品类描述','平台关键总成','本色件标识','是否精确追溯件','度量单位','是否法规件','备注']
                #df.columns = columns
                df_ect = df
                st.session_state['df_ect'] = df_ect
                
                #lines = df.values
                #lastCPAC = None
                #for line in lines:
                #    cpac = line[1]
                #    if pd.isna(cpac):
                #        line[1:4] = lastCPAC                  
                #    else:
                #        lastCPAC = line[1:4]   
                #df_ect = pd.DataFrame(lines)
                #df_sys.columns = columns
                #st.session_state['df_sys'] = df_sys
            #'ectDict:', ectDict
            st.success('工程配置表文件: %s 已上传' % fname)

        if df_sys.empty or df_ect.empty:
            if df_sys.empty:
                st.warning('请上传CPAC&FND查询表')
            if df_ect.empty:
                st.warning('请上传工程配置表')
        else:
            ectValues = list(df_ect.iloc[:, 0].astype(str).values) 
            minS = st.number_input('请输入单行最少打点数量', 1, 1000, 3)                     
            uploaded_lou = st.file_uploader("上传进行核对的LOU清单", type=["xlsx"])
            if uploaded_lou is not None:
                #df_lou = pd.read_excel(uploaded_lou, header=8, usecols=range(0,21)) 
                #df_lou.columns = ['序号','更改说明','查找编号','产品类型','车辆系列','CPAC','CPAC_CN','备注','零件号',
                #                  '零件名称','GPC','FND','FND_CN','FND_EN','用量','度量单位','零件成熟度',
                #                  '替换件零件号','是否成套替换','LOU用法','用法（代码）'] 
                df_lou_all = pd.read_excel(uploaded_lou, header=8)
                lines = df_lou_all.values
                gfcDict = {} 
                i = 0               
                for line in lines:
                    i += 1
                    l = list(line[21:116])
                    cnt = line[14]
                    cnts = ''
                    try:
                        if (type(cnt) is int) or (type(cnt) is float):
                            cnts = '%d' % cnt
                        else:
                            cnts = '%s' % cnt
                    except Exception as e:
                        cnts = 'Error'                   
                    k = '%s_%s_%s' % (line[10], line[11], cnts)
                    if k not in gfcDict:
                        gfcDict[k] = [(i, l)]
                    else:
                        gfcDict[k].append((i, l))
                        #i, k, gfcDict[k]
                keys = gfcDict.keys()               

                minSErrors = []
                for key in keys:
                    lvvs = gfcDict[key]
                    if len(lvvs) > 1:
                        errors = checkLVVS(lvvs)
                        ep = st.expander('错误：[%s] 重复打点' % key)
                        if errors:
                            for rs in errors:
                                idxs = []
                                col = -1
                                for r in rs:
                                    idxs.append(str(r[0]))
                                    col = r[1]
                                ep.write('第 %s 行 第 %s 列 重复打点' % (','.join(idxs), col))                        
                    for lvv in lvvs:
                        lidx = lvv[0]
                        lvs = lvv[1]
                        sCnt = countListS(lvs)
                        if sCnt < minS:
                            minSErrors.append('第%s行 打点数量 为 %d 个' % (lidx, sCnt))
                if minSErrors:
                    with st.expander('错误: 打点数量 小于 %d 个'%minS):
                        for err in minSErrors:
                            err

                #cc = len(list(df_lou_all.columns))
                #cc
                df_lou = pd.read_excel(uploaded_lou, header=8, usecols=[5,6,8,9,10,11,12,13,14,19]) 
                df_lou.columns = ['CPAC','CPAC_CN','零件号','零件名称','GPC','FND','FND_CN','FND_EN','用量','LOU用法'] 
                values = {'CPAC_CN': '/','FND_EN':'/','FND_CN':'/','LOU用法':'/'}
                df_lou.fillna(value=values,inplace=True) 
                lines = df_lou.values
                #'df_sys', df_sys                
                i = 0
                n = len(lines)
                bar = st.progress(float(i) / float(n))
                '*' * 100
                for line in lines:
                    cpac = line[0]                    
                    gpc = line[4]
                    fnd = line[5]
                    cnt = line[8]
            
                    fnd_cn = line[6].strip()
                    louStr_o = str(line[9])
                    #louStr = louStr_o.replace('（','(').replace('）',')').replace(' ','')
                    #louStr
                    louStr = louStr_o
                    ts = louStr.split('&')
                    for t in ts:
                        vs = t.split('|')
                        if len(vs) == 1:
                            v = vs[0].replace('(','').replace('（','').replace(')','').replace('）','')
                            vs = [v]
                        for v in vs:                                
                            if not ((('（' in v) and ('）' in v)) or (('(' in v) and (')' in v))) :
                                v = v.replace('（','').replace('）','').replace('(','').replace(')','')
                            #if v in ectDict:
                            #    #v
                            #    st.dataframe([ectDict[v]])
                            #else:
                            #    'Error: 配置字典无【%s】配置键' % v
                            
                    #if louStr_o in ectDict:
                    #    louStr_o
                    #    st.dataframe([ectDict[louStr_o]])
                    #else:
                    #    'Error: 配置字典无【%s】配置键' % louStr_o
                    lineError = False
                    louErrors = [] 
                
                    #if ' ' in louStr_o:
                    #    louErrors.append('含有非法字符：:red[空格]')
                    if '(' in louStr_o:
                        louErrors.append('含有非法字符：:red[左括号(英文)]')
                    if ')' in louStr_o:
                        louErrors.append('含有非法字符：:red[右括号(英文)]')                    
                    louList = []
                    #louStr = louStr_o.replace('(','').replace('（','').replace(')','').replace('）','').replace(' ','')
                    louStr = louStr_o
                    #'lou用法:', louStr_o
                    ss1 = louStr.split('&')
                    for s1 in ss1:
                        ss2 = s1.split('|')
                        #'ss2Len:', len(ss2), ss2
                        def tricS(s):
                            s2 = s
                            if ('(' in s2) and (')' not in s2):
                                s2 = s2.replace('(','')
                            if (')' in s2) and ('(' not in s2):
                                s2 = s2.replace(')','')
                            if ('（' in s2) and ('）' not in s2):
                                s2 = s2.replace('（','')
                            if ('）' in s2) and ('（' not in s2):
                                s2 = s2.replace('）','')
                            if (s2.find('(') == 0) or (s2.find('（')==0):
                                s2 = s2[1:-1]
                            return s2    
                        if len(ss2) == 1:
                            #s2 = ss2[0].replace('(','').replace('（','').replace(')','').replace('）','')
                            #'s2:',s2
                            s2 = ss2[0]
                            s2 = tricS(s2)
                            louList.append(s2)
                        else:    
                            for s2 in ss2:
                                #if not ((('（' in s2) and ('）' in s2)) or (('(' in s2) and (')' in s2))) :
                                #    s2 = s2.replace('（','').replace('）','').replace('(','').replace(')','')
                                #'s2:',s2
                                s2 = tricS(s2)
                                louList.append(s2)
                    #'louList', louList
                    for s in louList:
                        if s not in ectValues:
                            louErrors.append('不是合法的特征值：:red[%s]' % s)
                    if louErrors:
                        lineError = True
                        '第%d行 lou用法 :blue[%s] 中有如下错误：' % (i, louStr_o)
                        for err in louErrors:                       
                            st.markdown(err)

                    q = 'GPC == "%s" and FND == "%s"' % (gpc, fnd) 
                    if type(gpc) is int:
                        q = 'GPC == %s and FND == "%s"' % (gpc, fnd)
                    r = df_sys.query(q)                    
                    if len(r) == 0:
                        ':red[Error]: 第%d行 GPC: %s FND: %s 在系统表中不存在' % (i, gpc, fnd)
                        'LOU - [cpac] %s [fnd_cn] %s' % (cpac, fnd_cn)
                        lineError = True
                    else:
                        sys_line = r.values[0]
                        cpac_sys = sys_line[1]
                        fnd_cn_sys = sys_line[5]
                        cls = ['blue', 'blue']
                        if not (cpac == cpac_sys):
                            cls[0] = 'red'
                        if not (fnd_cn == fnd_cn_sys):
                            cls[1] = 'red'
                        if not (cpac == cpac_sys) or not (fnd_cn == fnd_cn_sys): 
                            st.write(':red[Error]: 第%d行 GPC: %s FND: %s' % (i, gpc, fnd))
                            'LOU - [cpac] :%s[%s] [fnd_cn] :%s[%s]' % (cls[0], cpac, cls[1], fnd_cn)
                            'SYS - [cpac] :%s[%s] [fnd_cn] :%s[%s]' % (cls[0], cpac_sys, cls[1], fnd_cn_sys)
                            lineError = True 
                    if lineError:
                        '*' * 100
                    i += 1
                    bar.progress(float(i) / float(n))
                #data = df_lou.to_csv(r'errors.csv', index=False, encoding='gbk')
                #st.download_button('下载错误列表', data, 'errors.csv')


    from openpyxl import load_workbook
    from openpyxl import Workbook
    def showCMANTool():
        st.header('CMAN统计工具')

        df = None
        cmanFile = ''
        if 'cmanFile' in st.session_state:
            cmanFile = st.session_state['cmanFile']
        if 'cmanDf' in st.session_state:
            df = st.session_state['cmanDf']            
        uploaded_file = st.file_uploader("上传进行CMAN统计的EBOM清单", type=["xlsx"])
        headRow = 8
        strikeRows = []
        if uploaded_file is not None:
            fname = uploaded_file.name
            if not (cmanFile == fname):
                st.session_state['cmanFile'] = fname
                wb=load_workbook(uploaded_file)
                sheet=wb.worksheets[0]
                for idx in range(1,sheet.max_row):                
                    c=sheet.cell(row=idx,column=3)
                    if c.value!=None:
                        if c.font.strike:
                            strikeRows.append(idx-headRow-2)
                            st.write('删除线数据行: %s  数据: %s' % (idx-headRow-1, c.value))
                    
                cmanOps = {}
                if 'cmanOps' in st.session_state: 
                    del st.session_state['cmanOps']
                df = pd.read_excel(uploaded_file, header=headRow)
                df = df.drop(index=strikeRows)      
                df['数量'] =  pd.to_numeric(df['数量'], errors='coerce')
                df = df[(df['数量'] > 0) & ((df['零部件来源']=='采购件') | (df['零部件来源']=='自制件') | (df['零部件来源']=='自制')) & (df['零部件标准件化']!='标准件')]
                st.session_state['cmanDf'] = df
            else:
                df = st.session_state['cmanDf']    

            st.write('添加筛选条件')            
            cs = list(df.columns)
            c = st.selectbox('选择筛选项',cs)
            v = st.selectbox('选择筛选值',df[c].unique())            
            if 'cmanOps' in st.session_state:
                cmanOps = st.session_state['cmanOps']
            if st.button('添加'):
                cmanOps[c] = v
            if st.button('删除'):
                if c in cmanOps:
                    del cmanOps[c]
            if st.button('清空'):
                del st.session_state['cmanOps']
                cmanOps = {}
            st.session_state['cmanOps'] = cmanOps
            st.write(cmanOps)
            
            dops = None
            cs = ''
            ops = None
            for c in cmanOps.keys():
                v = cmanOps[c]
                df = df[df[c] == v]
            st.write(df)

            #if ops:
            #    df["层级"] = df["层级"].astype('int')

            def getCMAN(dfm, dp):
                r = {}
                dfx = dfm
                if dp:
                    dfx = df[df['部门'] == dp]                
                df1 = dfx.groupby('CMAN',as_index=False).sum()[['CMAN','数量']]            
                vs = list(df1[df1['CMAN'] == 'C'].values)
                vc = 0 if (not vs) else vs[0][1]
                vs = list(df1[df1['CMAN'] == 'M'].values)           
                vm = 0 if (not vs) else vs[0][1]
                vs = list(df1[df1['CMAN'] == 'A'].values)           
                va = 0 if (not vs) else vs[0][1]
                vs = list(df1[df1['CMAN'] == 'N'].values)
                vn = 0 if (not vs) else vs[0][1]
                if not dp:
                    dp = '整车目标'
                vAll = vc+vm+va+vn
                vp = 0
                if vAll > 0:
                    vp = '%.2f%%' % ((vc+vm)/(vAll)*100)
                else:
                    '部门:%s vAll:%s' % (dp, vAll)
                r = {'部门':dp, '数量':vAll, 'C':vc, 'M':vm, 'A':va, 'N':vn, '实际':vp }
                #st.write(r)
                return r
                #st.write('总数：%s C：%s    M：%s    A：%s    N：%s    CMAN：%s  部门：%s ' % (vAll, vc, vm, va, vn, vp, dp))
                
            r = getCMAN(df, None)
            dps = df['部门'].unique()
            rs = [r]
            for dp in dps:
                r = getCMAN(df, dp)
                rs.append(r)
                
            st.dataframe(rs)
                      
            
            #st.write(df1)
            #st.write()
            
    def showCfgTool():
        uploaded_file = st.file_uploader("上传需处理的工程配置表", type=["xlsx"])
        headRow = 0
        rs = []
        if uploaded_file is not None:
            fname = uploaded_file.name
            df = pd.read_excel(uploaded_file, header=headRow) 
            df = df.fillna('-')
            df
            vs = list(df.values)
            rs.append(vs[0])
            rs.append(vs[1])
            vs = vs[2:]
            for arr in vs:
                a = arr[13:]
                aStr = ''.join(str(a)).upper()
                if 'S' in aStr:
                   rs.append(arr)
            df = pd.DataFrame(rs).set_index(0)
            st.dataframe(df) 
            import io
            buffer = io.BytesIO()
            fn = fname.split('.x')[0]
            with pd.ExcelWriter(buffer, engine='xlsxwriter') as writer:
                sheet_new = 'Sheet1'
                df.to_excel(writer, sheet_name=sheet_new, index=False)                 
                writer.close()
                fileName= '%s_处理后.xlsx' % (fn)
                st.download_button(
                    label="下载文件",
                    data=buffer,
                    file_name=fileName,
                    mime='application/vnd.ms-excel'
                    )
        

    def ShowBomDiffDatabase():
        boms = {}
        bomFile = ''
        diffBomDb = ''
        if 'bomFile' in st.session_state:
            bomFile = st.session_state['bomFile']
        if 'bomData' in st.session_state:
            boms = st.session_state['bomData']
        if 'diffBomDb' in st.session_state:
            diffBomDb = st.session_state['diffBomDb']    
        
        st.header('BOM表单与数据库中数据进行核对')

        uploaded_file = st.file_uploader("上传需要核对的EBOM清单", type=["xlsx"])
        if uploaded_file is not None:
            fname = uploaded_file.name
            if not (bomFile == fname):
                st.session_state['bomFile'] = fname
                df = pd.read_excel(uploaded_file, sheet_name=None)
                boms = {}
                loadBar = st.progress(0, text='')
                for sheetName in df.keys():
                    try:
                        df = pd.read_excel(uploaded_file, header=9, usecols=['零件编号','零件名称','数量','层级'], sheet_name=sheetName)
                        bomArray = df.values
                        systems = {}
                        system = None
                        n = len(bomArray)
                        i = 0
                        pText = '加载Ebom[%s] (%d/%d)'%(sheetName, i, n)
                        for bomLine in bomArray:
                            sysId = bomLine[0]
                            sysLevel = bomLine[3]
                            if sysLevel == '2Y':
                                system = {'sys':bomLine , 'children':[]}
                                systems[sysId] = system
                            else:
                                if system:
                                    system['children'].append(bomLine)
                            i += 1
                            pText = '加载Ebom[%s] (%d/%d) 完成！'%(sheetName, i, n)
                            loadBar.progress(float(i)/float(n), text='')
                        st.write(pText)
                        boms[sheetName] = systems
                    except:
                        #pText = '加载Ebom[%s]失败，格式不正确'%sheetName
                        #st.write(pText)
                        pass
                    finally:
                        st.session_state['bomData'] = boms
                #st.write(boms)
            bomKeys = list(boms.keys())
            st.subheader('通过对比Bom数据库核查问题')
            bk = st.selectbox('选择一个BOM：',bomKeys)
            st.session_state['diffBomDb'] = bk
            bom = boms[bk]            
            sns = bom.keys()
            cols = st.columns(2)
            cols[0].header(bk)
            cols[1].header('BOM数据库')
            loadBar = st.progress(0, text='')
            ep1 = st.expander("2Y")
            ep2 = st.expander("2Y子件")        
            cols1 = ep1.columns(2)
            cols2 = ep2.columns(2)
            i = 0
            n = len(sns)
            for sn in sns:
                ss = bom[sn]                
                SystemChildrenDiff_DB(ss, cols1, cols2, ep2)
                i += 1
                pText = '核对进度： (%d/%d)'%(i, n)
                loadBar.progress(float(i)/float(n), text=pText)
            loadBar.progress(1.0, text='Bom清单【:blue[%s]】与数据库比对结束!' % bk)    
                
    if s == 'BOM数据库维护':
        ShowBomDatabaseApp()
    elif s == '数据库问题核查':
        ShowBomDatabaseProblems()
    elif s == 'BOM数据表核对':
        ShowBomDiffSheets()
    elif s == 'BOM数据库核对':
        ShowBomDiffDatabase()
    elif s =='CMAN统计工具':
        showCMANTool()
    elif s =='配置表处理工具':
        showCfgTool()
    elif s == '差异件清单生成':
        showDiffSheetTool()
    elif s == 'MBOM差异件清单生成':
        showMBOMDiffSheetTool()
    elif s == 'EBOM&MBOM差异件清单生成':
        showEBOMAndMBOMDiffSheetTool()
    elif s == 'LOU核查工具':
        showLOUTool()
    elif s == 'LOU打点工具':
        showLOUPointTool()
    elif s == '工程配置工具':
        showECTTool()
    elif s == 'BOM差异件核对':
        showBOMDiffTool()
    elif s == '配置打点信息一致性核查':
        showLOUCheckTool()
    else:
        st.title('其他BOM工具持续开发中，敬请期待！')

bomTools()

